from __future__ import annotations

import io
import unicodedata
from datetime import date, datetime, timedelta
from decimal import ROUND_CEILING, ROUND_HALF_UP, Decimal

from django.db import transaction
from django.db.models import Case, Count, DecimalField, F, Q, Sum, Value, When
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from rest_framework import status, viewsets
from rest_framework.authtoken.models import Token
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import AllowAny, BasePermission, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.teams import (
    BalanceablePlayer,
    GreedyTeamBalancer,
    TeamBalanceConfig,
    TeamGenerationRequest,
)

from .models import (
    Match,
    MatchAttendance,
    MatchPlayerRating,
    MatchPlayerStat,
    Player,
    Role,
    Transaction,
    User,
)
from .permissions import IsAdminOrReadOnly
from .serializers import (
    AdminResetPasswordSerializer,
    ChangePasswordSerializer,
    FinancialSummarySerializer,
    LoginSerializer,
    MatchAttendanceSerializer,
    MatchFinalizeSerializer,
    MatchPlayerRatingStateSerializer,
    MatchPlayerRatingSubmitSerializer,
    MatchSerializer,
    MatchStatsImportSummarySerializer,
    PaymentRankingSerializer,
    PlayerSerializer,
    PortalOverviewSerializer,
    PresenceRankingSerializer,
    PublicSignupSerializer,
    SeasonOverviewSerializer,
    SportsRankingSerializer,
    TeamGenerationInputSerializer,
    TeamGenerationResponseSerializer,
    TeamPlayerSwapSerializer,
    TransactionSerializer,
    UserAccountSerializer,
    UserProfileUpdateSerializer,
    UserSerializer,
)

GUEST_FEE_AMOUNT = Decimal("14.00")
FINAL_MATCH_STATUSES = [Match.Status.ARCHIVED]
UNLINKED_RATING_LOCK_REASON = (
    "Sua conta ainda nao esta vinculada a um jogador. "
    "Procure um administrador para validar sua conta."
)
MISSING_TEAM_RATING_LOCK_REASON = (
    "Seu jogador vinculado nao possui time gerado nesta pelada. "
    "A votacao fica disponivel apenas para jogadores do mesmo time."
)


def month_bounds(reference_date: date) -> tuple[date, date]:
    month_start = reference_date.replace(day=1)
    if month_start.month == 12:
        month_end = date(month_start.year + 1, 1, 1)
    else:
        month_end = date(month_start.year, month_start.month + 1, 1)
    return month_start, month_end


def monthly_reference_filter(month_start: date, month_end: date) -> Q:
    return Q(reference_month__gte=month_start, reference_month__lt=month_end) | (
        Q(reference_month__isnull=True)
        & Q(occurred_on__gte=month_start)
        & Q(occurred_on__lt=month_end)
    )


def parse_reference_month(raw_reference_month: str | None) -> tuple[date, str]:
    if not raw_reference_month:
        today = timezone.localdate()
        return today.replace(day=1), today.strftime("%Y-%m")

    try:
        parsed = datetime.strptime(raw_reference_month, "%Y-%m").date().replace(day=1)
    except ValueError as exc:
        raise ValidationError({"reference_month": "Use o formato YYYY-MM."}) from exc
    return parsed, parsed.strftime("%Y-%m")


def build_player_payment_map(player_ids: list, month_start: date, month_end: date) -> dict:
    month_transactions = (
        Transaction.objects.filter(
            related_player_id__in=player_ids,
            category=Transaction.Category.MONTHLY_FEE,
            direction=Transaction.Direction.INFLOW,
        )
        .filter(monthly_reference_filter(month_start, month_end))
        .values("related_player_id")
        .annotate(
            paid_amount=Coalesce(
                Sum(
                    Case(
                        When(status=Transaction.Status.POSTED, then=F("amount")),
                        default=Value(0),
                        output_field=DecimalField(max_digits=10, decimal_places=2),
                    )
                ),
                Decimal("0.00"),
            ),
            pending_amount=Coalesce(
                Sum(
                    Case(
                        When(status=Transaction.Status.PENDING, then=F("amount")),
                        default=Value(0),
                        output_field=DecimalField(max_digits=10, decimal_places=2),
                    )
                ),
                Decimal("0.00"),
            ),
        )
    )
    return {
        row["related_player_id"]: {
            "paid_amount": row["paid_amount"],
            "pending_amount": row["pending_amount"],
        }
        for row in month_transactions
    }


def guest_fee_due_queryset():
    return MatchAttendance.objects.filter(
        is_guest=True,
        attendance_status=MatchAttendance.AttendanceStatus.CONFIRMED,
        guest_fee_status=MatchAttendance.GuestFeeStatus.PENDING,
        guest_fee_amount__gt=Decimal("0.00"),
        match__status__in=FINAL_MATCH_STATUSES,
    )


def get_guest_fee_pending_total():
    return guest_fee_due_queryset().aggregate(
        total=Coalesce(
            Sum("guest_fee_amount"),
            Decimal("0.00"),
            output_field=DecimalField(max_digits=12, decimal_places=2),
        )
    )["total"]


def clamp_overall(value: int) -> int:
    return max(0, min(99, value))


RATING_WINDOW_DURATION = timedelta(hours=24)
RATING_OVERALL_WEIGHT = Decimal("0.25")
OVERALL_HISTORY_START_DATE = date(2026, 5, 26)
OVERALL_HISTORY_EXCLUDED_DATES = [date(2026, 6, 30)]


def get_rating_performance_adjustment(rating_average: Decimal) -> Decimal:
    if rating_average >= Decimal("9.0"):
        return Decimal("2")

    if rating_average >= Decimal("8.0"):
        return Decimal("1")

    if rating_average <= Decimal("2.0"):
        return Decimal("-2")

    if rating_average <= Decimal("4.0"):
        return Decimal("-1")

    return Decimal("0")


def get_rating_average(scores: list[Decimal]) -> Decimal | None:
    if not scores:
        return None

    ratings = [Decimal(str(score)) for score in scores]
    return sum(ratings, Decimal("0")) / Decimal(len(ratings))


def get_final_rating_score(rating_average: Decimal) -> Decimal:
    return rating_average.quantize(Decimal("1"), rounding=ROUND_CEILING)


def get_match_rating_stats(match: Match) -> dict:
    rating_stats = {}
    for rating in MatchPlayerRating.objects.filter(match=match).values(
        "rated_attendance_id",
        "score",
    ):
        stats = rating_stats.setdefault(
            rating["rated_attendance_id"],
            {"scores": [], "rating_count": 0},
        )
        stats["scores"].append(rating["score"])
        stats["rating_count"] += 1

    for stats in rating_stats.values():
        stats["average_score"] = get_rating_average(stats["scores"])
        del stats["scores"]

    return rating_stats


def recalculate_player_overall_from_match_ratings(
    player: Player, match: Match, *, base_overall: int | None = None
) -> None:
    rating_average = get_rating_average(
        list(
            MatchPlayerRating.objects.filter(match=match, rated_player=player).values_list(
                "score", flat=True
            )
        )
    )
    if rating_average is None:
        return

    current_overall = player.overall if base_overall is None else base_overall
    rating_average_decimal = Decimal(str(rating_average))
    final_rating_score = get_final_rating_score(rating_average_decimal)
    community_overall = final_rating_score * Decimal("10")
    base_delta = (community_overall - Decimal(current_overall)) * RATING_OVERALL_WEIGHT
    performance_adjustment = get_rating_performance_adjustment(final_rating_score)
    overall_delta = (base_delta + performance_adjustment).quantize(
        Decimal("1"), rounding=ROUND_HALF_UP
    )
    next_overall = Decimal(current_overall) + overall_delta
    player.overall = clamp_overall(int(next_overall))
    player.save(update_fields=["overall", "updated_at"])


def recalculate_finalized_match_ratings(match: Match) -> bool:
    if match.status != Match.Status.ARCHIVED or not match.ratings_finalized_at:
        return False

    with transaction.atomic():
        locked_match = Match.objects.select_for_update().get(pk=match.pk)
        if locked_match.status != Match.Status.ARCHIVED or not locked_match.ratings_finalized_at:
            match.ratings_finalized_at = locked_match.ratings_finalized_at
            return False

        rated_attendance_ids = (
            MatchPlayerRating.objects.filter(match=locked_match)
            .values_list("rated_attendance_id", flat=True)
            .distinct()
        )
        attendance_overall_by_player_id = {
            entry.player_id: entry.overall
            for entry in MatchAttendance.objects.filter(
                match=locked_match,
                id__in=rated_attendance_ids,
                player__isnull=False,
            )
        }

        for player in Player.objects.select_for_update().filter(
            id__in=attendance_overall_by_player_id
        ):
            recalculate_player_overall_from_match_ratings(
                player,
                locked_match,
                base_overall=attendance_overall_by_player_id[player.id],
            )

        match.ratings_finalized_at = locked_match.ratings_finalized_at
        return True


def get_rating_window_started_at(match: Match):
    if match.status != Match.Status.ARCHIVED:
        return None
    return match.archived_at or match.updated_at


def get_rating_window_closes_at(match: Match):
    started_at = get_rating_window_started_at(match)
    if started_at is None:
        return None
    return started_at + RATING_WINDOW_DURATION


def is_rating_window_expired(match: Match) -> bool:
    window_closes_at = get_rating_window_closes_at(match)
    return bool(window_closes_at and timezone.now() >= window_closes_at)


def finalize_match_ratings(match: Match, *, force: bool = False) -> bool:
    if (
        match.status != Match.Status.ARCHIVED
        or match.ratings_finalized_at
        or (not force and not is_rating_window_expired(match))
    ):
        return False

    with transaction.atomic():
        locked_match = Match.objects.select_for_update().get(pk=match.pk)
        if (
            locked_match.status != Match.Status.ARCHIVED
            or locked_match.ratings_finalized_at
            or (not force and not is_rating_window_expired(locked_match))
        ):
            match.ratings_finalized_at = locked_match.ratings_finalized_at
            return False

        rated_player_ids = (
            MatchPlayerRating.objects.filter(match=locked_match)
            .values_list("rated_player_id", flat=True)
            .distinct()
        )
        for player in Player.objects.filter(id__in=rated_player_ids):
            recalculate_player_overall_from_match_ratings(player, locked_match)

        locked_match.ratings_finalized_at = timezone.now()
        locked_match.save(update_fields=["ratings_finalized_at", "updated_at"])
        match.ratings_finalized_at = locked_match.ratings_finalized_at
        return True


def finalize_match_ratings_if_due(match: Match) -> bool:
    return finalize_match_ratings(match, force=False)


def finalize_due_match_ratings() -> None:
    due_matches = Match.objects.filter(
        status=Match.Status.ARCHIVED,
        ratings_finalized_at__isnull=True,
    )
    for match in due_matches:
        finalize_match_ratings_if_due(match)


MATCH_FINALIZE_SOURCE_LABEL = "Finalizacao da pelada"


def get_confirmed_attendance_entries(match: Match) -> list[MatchAttendance]:
    return list(
        match.attendance_entries.select_related("player").filter(
            attendance_status=MatchAttendance.AttendanceStatus.CONFIRMED,
        )
    )


def get_assigned_team_numbers(attendance_entries) -> list[int]:
    return sorted(
        {
            entry.assigned_team_number
            for entry in attendance_entries
            if entry.assigned_team_number is not None
        }
    )


def sync_match_win_stats(
    match: Match,
    attendance_entries,
    winning_team_number: int,
    user: User | None,
) -> None:
    """Mark the winning squad on the per-player stats that feed the wins ranking."""

    existing_stats = {
        stat.attendance_id: stat for stat in MatchPlayerStat.objects.filter(match=match)
    }
    stats_to_create = []
    stats_to_update = []

    for entry in attendance_entries:
        if entry.assigned_team_number is None:
            continue

        team_won = entry.assigned_team_number == winning_team_number
        stat = existing_stats.get(entry.id)
        if stat is None:
            stats_to_create.append(
                MatchPlayerStat(
                    match=match,
                    attendance=entry,
                    player=entry.player,
                    display_name=entry.display_name,
                    team_number=entry.assigned_team_number,
                    team_name=entry.assigned_team_name,
                    goals=0,
                    assists=0,
                    team_won=team_won,
                    imported_by=user if user and user.is_authenticated else None,
                    source_label=MATCH_FINALIZE_SOURCE_LABEL,
                )
            )
        elif stat.team_won != team_won:
            stat.team_won = team_won
            stats_to_update.append(stat)

    if stats_to_create:
        MatchPlayerStat.objects.bulk_create(stats_to_create)
    if stats_to_update:
        MatchPlayerStat.objects.bulk_update(stats_to_update, ["team_won", "updated_at"])


def finalize_match(match: Match, winning_team_number: int | None, user: User | None) -> Match:
    attendance_entries = get_confirmed_attendance_entries(match)
    team_numbers = get_assigned_team_numbers(attendance_entries)

    if winning_team_number is not None:
        if len(team_numbers) != 2:
            raise ValidationError(
                {
                    "winning_team_number": (
                        "So e possivel escolher o time vencedor em peladas com dois times."
                    )
                }
            )
        if winning_team_number not in team_numbers:
            raise ValidationError({"winning_team_number": "Escolha um dos times desta pelada."})

    now = timezone.now()
    with transaction.atomic():
        match.status = Match.Status.ARCHIVED
        match.attendance_locked_at = match.attendance_locked_at or now
        match.archived_at = match.archived_at or now
        update_fields = ["status", "attendance_locked_at", "archived_at", "updated_at"]

        if winning_team_number is not None:
            match.winning_team_number = winning_team_number
            update_fields.append("winning_team_number")
            sync_match_win_stats(match, attendance_entries, winning_team_number, user)

        match.save(update_fields=update_fields)

    return match


MATCH_STATS_SHEET_NAME = "Estatisticas"
MATCH_STATS_HEADERS = ["Jogador", "Perfil", "Gols", "Assistencias"]
TRUTHY_SHEET_VALUES = {
    "1",
    "x",
    "s",
    "sim",
    "true",
    "verdadeiro",
    "vitoria",
    "ganhou",
    "win",
    "yes",
}
XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def normalize_csv_key(value: str | None) -> str:
    normalized = unicodedata.normalize("NFKD", value or "")
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    return "".join(character for character in ascii_value.lower() if character.isalnum())


def normalize_sheet_value(value: object) -> str:
    return str(value or "").strip()


def parse_sheet_integer(value: object, field_name: str) -> int:
    if not value:
        return 0
    try:
        parsed = int(Decimal(str(value).replace(",", ".")))
    except Exception as exc:
        raise ValidationError({field_name: f"Valor invalido: {value}."}) from exc
    if parsed < 0:
        raise ValidationError({field_name: "Use zero ou um numero positivo."})
    return parsed


def parse_sheet_boolean(value: object) -> bool:
    return normalize_csv_key(str(value or "")) in TRUTHY_SHEET_VALUES


def get_attendance_team_name(attendance: MatchAttendance) -> str:
    return attendance.assigned_team_name or (
        f"Time {attendance.assigned_team_number}"
        if attendance.assigned_team_number is not None
        else "Sem time"
    )


def get_attendance_visible_team_key(attendance: MatchAttendance) -> tuple[str, str]:
    return ("name", normalize_csv_key(get_attendance_team_name(attendance)))


def build_match_stats_xlsx_response(match: Match) -> HttpResponse:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = MATCH_STATS_SHEET_NAME

    title_fill = PatternFill("solid", fgColor="312E81")
    team_fill = PatternFill("solid", fgColor="EDE9FE")
    header_fill = PatternFill("solid", fgColor="4F46E5")
    editable_fill = PatternFill("solid", fgColor="FEF3C7")
    border_side = Side(style="thin", color="D4D4D8")
    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    center_alignment = Alignment(horizontal="center", vertical="center")

    worksheet.merge_cells("A1:D1")
    worksheet["A1"] = (
        f"Pelada de {timezone.localtime(match.scheduled_at).strftime('%d/%m/%Y %H:%M')}"
    )
    worksheet["A1"].fill = title_fill
    worksheet["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    worksheet["A1"].alignment = center_alignment
    worksheet.row_dimensions[1].height = 28

    worksheet.merge_cells("A2:D2")
    worksheet["A2"] = match.location or "Local a definir"
    worksheet["A2"].font = Font(color="52525B", italic=True)
    worksheet["A2"].alignment = center_alignment
    worksheet.row_dimensions[2].height = 22

    worksheet.column_dimensions["A"].width = 32
    worksheet.column_dimensions["B"].width = 16
    worksheet.column_dimensions["C"].width = 14
    worksheet.column_dimensions["D"].width = 18

    attendance_entries = list(
        match.attendance_entries.select_related("player")
        .filter(attendance_status=MatchAttendance.AttendanceStatus.CONFIRMED)
        .order_by("assigned_team_number", "assigned_team_name", "display_name")
    )
    stats_by_attendance_id = {
        stat.attendance_id: stat for stat in MatchPlayerStat.objects.filter(match=match)
    }
    team_numbers = sorted(
        {
            entry.assigned_team_number
            for entry in attendance_entries
            if entry.assigned_team_number is not None
        }
    )

    current_row = 4
    for team_number in team_numbers:
        team_entries = [
            entry for entry in attendance_entries if entry.assigned_team_number == team_number
        ]
        team_name = team_entries[0].assigned_team_name or f"Time {team_number}"
        team_won = any(
            stats_by_attendance_id.get(entry.id) and stats_by_attendance_id[entry.id].team_won
            for entry in team_entries
        )
        worksheet.merge_cells(
            start_row=current_row, start_column=1, end_row=current_row, end_column=2
        )
        worksheet.cell(current_row, 1, team_name)
        worksheet.cell(current_row, 3, "Vitoria do time")
        worksheet.cell(current_row, 4, "SIM" if team_won else "")
        for column in range(1, 5):
            cell = worksheet.cell(current_row, column)
            cell.fill = team_fill if column < 4 else editable_fill
            cell.font = Font(bold=True, color="312E81")
            cell.border = cell_border
            cell.alignment = center_alignment if column >= 3 else Alignment(vertical="center")
        worksheet.row_dimensions[current_row].height = 24
        current_row += 1

        for index, header in enumerate(MATCH_STATS_HEADERS, start=1):
            cell = worksheet.cell(current_row, index, header)
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.border = cell_border
            cell.alignment = center_alignment
        worksheet.row_dimensions[current_row].height = 24
        current_row += 1

        for entry in team_entries:
            stat = stats_by_attendance_id.get(entry.id)
            values = [
                entry.display_name,
                "Convidado" if entry.is_guest else "Mensalista",
                stat.goals if stat else "",
                stat.assists if stat else "",
            ]
            for index, value in enumerate(values, start=1):
                cell = worksheet.cell(current_row, index, value)
                cell.border = cell_border
                cell.alignment = (
                    center_alignment if index in {2, 3, 4} else Alignment(vertical="center")
                )
                if index in {3, 4}:
                    cell.fill = editable_fill
                    cell.protection = Protection(locked=False)
            worksheet.row_dimensions[current_row].height = 24
            current_row += 1

        current_row += 1

    unassigned_entries = [
        entry for entry in attendance_entries if entry.assigned_team_number is None
    ]
    if unassigned_entries:
        worksheet.merge_cells(
            start_row=current_row, start_column=1, end_row=current_row, end_column=4
        )
        worksheet.cell(current_row, 1, "Sem time")
        for column in range(1, 5):
            cell = worksheet.cell(current_row, column)
            cell.fill = team_fill
            cell.font = Font(bold=True, color="312E81")
            cell.border = cell_border
        worksheet.row_dimensions[current_row].height = 24
        current_row += 1

        for index, header in enumerate(MATCH_STATS_HEADERS, start=1):
            cell = worksheet.cell(current_row, index, header)
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.border = cell_border
            cell.alignment = center_alignment
        worksheet.row_dimensions[current_row].height = 24
        current_row += 1

        for entry in unassigned_entries:
            stat = stats_by_attendance_id.get(entry.id)
            values = [
                entry.display_name,
                "Convidado" if entry.is_guest else "Mensalista",
                stat.goals if stat else "",
                stat.assists if stat else "",
            ]
            for index, value in enumerate(values, start=1):
                cell = worksheet.cell(current_row, index, value)
                cell.border = cell_border
                cell.alignment = (
                    center_alignment if index in {2, 3, 4} else Alignment(vertical="center")
                )
                if index in {3, 4}:
                    cell.fill = editable_fill
                    cell.protection = Protection(locked=False)
            worksheet.row_dimensions[current_row].height = 24
            current_row += 1

    for row in worksheet.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=False)
    worksheet.freeze_panes = "A4"
    worksheet.sheet_view.showGridLines = False
    worksheet.protection.sheet = False

    buffer = io.BytesIO()
    workbook.save(buffer)
    filename = f"estatisticas-pelada-{match.scheduled_at.date().isoformat()}.xlsx"
    response = HttpResponse(buffer.getvalue(), content_type=XLSX_CONTENT_TYPE)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def import_match_stats_from_xlsx(match: Match, uploaded_file, user: User) -> dict:
    try:
        workbook = load_workbook(uploaded_file, data_only=True)
    except Exception as exc:
        raise ValidationError({"file": "Envie a planilha XLSX exportada pelo sistema."}) from exc

    if MATCH_STATS_SHEET_NAME not in workbook.sheetnames:
        raise ValidationError({"file": "A planilha precisa ser o XLSX exportado pelo sistema."})

    worksheet = workbook[MATCH_STATS_SHEET_NAME]
    attendance_entries = list(
        match.attendance_entries.select_related("player").filter(
            attendance_status=MatchAttendance.AttendanceStatus.CONFIRMED
        )
    )
    attendance_by_visible_key = {}
    duplicate_visible_keys = set()
    for entry in attendance_entries:
        visible_key = (
            get_attendance_visible_team_key(entry),
            normalize_csv_key(entry.display_name),
        )
        if visible_key in attendance_by_visible_key:
            duplicate_visible_keys.add(visible_key)
            continue
        attendance_by_visible_key[visible_key] = entry

    parsed_stats_by_attendance_id = {}
    winning_team_keys = set()
    player_rows_found = 0
    current_team_key = None

    for row_number in range(4, worksheet.max_row + 1):
        first_value = normalize_sheet_value(worksheet.cell(row_number, 1).value)
        second_value = normalize_sheet_value(worksheet.cell(row_number, 2).value)
        third_value = normalize_sheet_value(worksheet.cell(row_number, 3).value)
        fourth_value = worksheet.cell(row_number, 4).value
        normalized_first = normalize_csv_key(first_value)
        normalized_third = normalize_csv_key(third_value)

        if not any([first_value, second_value, third_value, fourth_value]):
            continue

        if normalized_first == "jogador":
            continue

        is_team_row = bool(first_value) and (
            normalized_third == "vitoriadotime" or (not second_value and not third_value)
        )
        if is_team_row:
            current_team_key = ("name", normalize_csv_key(first_value))
            if parse_sheet_boolean(fourth_value):
                winning_team_keys.add(current_team_key)
            continue

        if not first_value or current_team_key is None:
            continue

        visible_key = (current_team_key, normalize_csv_key(first_value))
        if visible_key in duplicate_visible_keys:
            raise ValidationError(
                {
                    "file": f"Ha jogadores com nome duplicado no time {first_value}. Renomeie antes de importar."
                }
            )

        attendance_entry = attendance_by_visible_key.get(visible_key)
        if not attendance_entry:
            raise ValidationError({"file": f"Jogador nao encontrado nesta pelada: {first_value}."})
        attendance_id = str(attendance_entry.id)
        if attendance_id in parsed_stats_by_attendance_id:
            raise ValidationError({"file": f"Jogador duplicado na planilha: {first_value}."})

        player_rows_found += 1
        parsed_stats_by_attendance_id[attendance_id] = {
            "goals": parse_sheet_integer(worksheet.cell(row_number, 3).value, "gols"),
            "assists": parse_sheet_integer(worksheet.cell(row_number, 4).value, "assistencias"),
        }

    if player_rows_found == 0:
        raise ValidationError({"file": "Nenhuma linha de jogador foi encontrada na planilha."})

    replaced_existing = MatchPlayerStat.objects.filter(match=match).count()
    # A planilha sem coluna de vitoria preenchida nao pode apagar o vencedor
    # escolhido ao finalizar a pelada.
    fallback_winning_team_number = match.winning_team_number if not winning_team_keys else None
    stats_to_create = []
    for attendance_entry in attendance_entries:
        imported_values = parsed_stats_by_attendance_id.get(
            str(attendance_entry.id),
            {"goals": 0, "assists": 0},
        )
        team_key = get_attendance_visible_team_key(attendance_entry)
        if fallback_winning_team_number is None:
            team_won = bool(team_key and team_key in winning_team_keys)
        else:
            team_won = attendance_entry.assigned_team_number == fallback_winning_team_number
        stats_to_create.append(
            MatchPlayerStat(
                match=match,
                attendance=attendance_entry,
                player=attendance_entry.player,
                display_name=attendance_entry.display_name,
                team_number=attendance_entry.assigned_team_number,
                team_name=attendance_entry.assigned_team_name,
                goals=imported_values["goals"],
                assists=imported_values["assists"],
                team_won=team_won,
                imported_by=user,
                source_label=getattr(uploaded_file, "name", "")[:120],
            )
        )

    winning_team_numbers = {
        stat.team_number
        for stat in stats_to_create
        if stat.team_won and stat.team_number is not None
    }

    with transaction.atomic():
        MatchPlayerStat.objects.filter(match=match).delete()
        MatchPlayerStat.objects.bulk_create(stats_to_create)
        if len(winning_team_numbers) == 1:
            single_winner = next(iter(winning_team_numbers))
            if match.winning_team_number != single_winner:
                match.winning_team_number = single_winner
                match.save(update_fields=["winning_team_number", "updated_at"])

    return {
        "match_id": match.id,
        "players_processed": len(stats_to_create),
        "goals_total": sum(stat.goals for stat in stats_to_create),
        "assists_total": sum(stat.assists for stat in stats_to_create),
        "winning_teams": sorted(
            {
                stat.team_name or f"Time {stat.team_number}"
                for stat in stats_to_create
                if stat.team_won
            }
        ),
        "replaced_existing": replaced_existing,
    }


def build_sports_ranking(limit: int) -> dict:
    aggregated = {}
    for stat in MatchPlayerStat.objects.select_related("player"):
        key = stat.player_id or f"guest:{normalize_csv_key(stat.display_name)}"
        entry = aggregated.setdefault(
            key,
            {
                "player_id": stat.player_id,
                "player_name": stat.player.full_name if stat.player_id else stat.display_name,
                "goals": 0,
                "assists": 0,
                "wins": 0,
            },
        )
        entry["goals"] += stat.goals
        entry["assists"] += stat.assists
        entry["wins"] += 1 if stat.team_won else 0

    entries = list(aggregated.values())

    def top_by(primary: str, *secondary: str):
        return [
            entry
            for entry in sorted(
                entries,
                key=lambda item: (
                    -item[primary],
                    *(-item[field] for field in secondary),
                    item["player_name"].lower(),
                ),
            )
            if entry[primary] > 0
        ][:limit]

    return {
        "top_scorers": top_by("goals", "assists", "wins"),
        "top_assistants": top_by("assists", "goals", "wins"),
        "top_winners": top_by("wins", "goals", "assists"),
    }


class IsRoleAdmin(BasePermission):
    def has_permission(self, request, view):
        return bool(
            request.user and request.user.is_authenticated and request.user.role == Role.ADMIN
        )


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.select_related("linked_player").all().order_by("username")
    serializer_class = UserAccountSerializer
    permission_classes = [IsRoleAdmin]

    @action(detail=True, methods=["post"], url_path="reset-password")
    def reset_password(self, request, pk=None):
        user = self.get_object()
        serializer = AdminResetPasswordSerializer(
            data=request.data,
            context={"request": request, "target_user": user},
        )
        serializer.is_valid(raise_exception=True)

        user.set_password(serializer.validated_data["new_password"])
        user.must_change_password = True
        user.save()
        Token.objects.filter(user=user).delete()
        return Response(self.get_serializer(user).data)


class PlayerViewSet(viewsets.ModelViewSet):
    queryset = Player.objects.all()
    serializer_class = PlayerSerializer
    permission_classes = [IsAdminOrReadOnly]

    def get_queryset(self):
        finalize_due_match_ratings()
        queryset = super().get_queryset()
        if getattr(self.request.user, "role", None) == Role.ADMIN:
            return queryset
        return queryset.filter(player_type=Player.PlayerType.MEMBER, is_active=True)


class MatchViewSet(viewsets.ModelViewSet):
    queryset = Match.objects.select_related("created_by").all()
    serializer_class = MatchSerializer
    permission_classes = [IsAdminOrReadOnly]

    def get_permissions(self):
        if self.action in {"export_stats_sheet", "import_stats_sheet"}:
            return [IsRoleAdmin()]
        if self.action == "player_ratings":
            return [IsAuthenticated()]
        return super().get_permissions()

    def get_queryset(self):
        finalize_due_match_ratings()
        queryset = super().get_queryset()
        if getattr(self.request.user, "role", None) == Role.ADMIN:
            return queryset
        return queryset.filter(status__in=[Match.Status.OPEN, Match.Status.ARCHIVED]).distinct()

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def perform_update(self, serializer):
        validated_data = serializer.validated_data
        status_value = validated_data.get("status", serializer.instance.status)
        result_summary = validated_data.get("result_summary", serializer.instance.result_summary)

        extra_fields = {}
        if "status" in validated_data:
            archived_at = None
            if status_value == Match.Status.ARCHIVED:
                archived_at = serializer.instance.archived_at or timezone.now()

            extra_fields["attendance_locked_at"] = (
                timezone.now() if status_value == Match.Status.ARCHIVED else None
            )
            extra_fields["archived_at"] = archived_at
            if status_value != Match.Status.ARCHIVED:
                extra_fields["ratings_finalized_at"] = None

        if "result_summary" in validated_data:
            extra_fields["result_recorded_at"] = timezone.now() if result_summary else None

        serializer.save(**extra_fields)

    @action(detail=False, methods=["get"], url_path="current")
    def current(self, request):
        match = (
            self.get_queryset().filter(status=Match.Status.OPEN).order_by("scheduled_at").first()
        )
        if match is None:
            return Response(status=status.HTTP_204_NO_CONTENT)
        return Response(self.get_serializer(match).data)

    @action(detail=True, methods=["post"], url_path="generate-teams")
    def generate_teams(self, request, pk=None):
        input_serializer = TeamGenerationInputSerializer(data=request.data)
        input_serializer.is_valid(raise_exception=True)
        match = self.get_object()
        team_count = input_serializer.validated_data.get("team_count", match.expected_team_count)
        attendance_entries = list(
            match.attendance_entries.filter(
                attendance_status=MatchAttendance.AttendanceStatus.CONFIRMED
            )
        )
        balancer = GreedyTeamBalancer()
        players = tuple(
            BalanceablePlayer(
                id=str(entry.id),
                display_name=entry.display_name,
                overall=entry.overall,
                preferred_position=(
                    entry.player.preferred_position if entry.player_id else "UNIVERSAL"
                ),
                is_guest=entry.is_guest,
            )
            for entry in attendance_entries
        )
        result = balancer.generate(
            TeamGenerationRequest(
                players=players,
                config=TeamBalanceConfig(team_count=team_count),
                match_id=match.id,
            )
        )

        team_payload = []
        with transaction.atomic():
            match.attendance_entries.update(
                assigned_team_number=None,
                assigned_team_name="",
            )
            for team_index, team in enumerate(result.teams, start=1):
                for player in team.players:
                    MatchAttendance.objects.filter(id=player.id).update(
                        assigned_team_number=team_index,
                        assigned_team_name=team.name,
                    )
                team_payload.append(
                    {
                        "name": team.name,
                        "total_overall": team.total_overall,
                        "average_overall": team.average_overall.quantize(Decimal("0.01")),
                        "players": [
                            {
                                "id": str(player.id),
                                "display_name": player.display_name,
                                "is_guest": player.is_guest,
                                "overall": player.overall,
                            }
                            for player in team.players
                        ],
                    }
                )

            match.teams_generated_at = timezone.now()
            match.expected_team_count = team_count
            match.save(update_fields=["teams_generated_at", "expected_team_count", "updated_at"])

        payload = {
            "match_id": match.id,
            "average_overall_gap": result.average_overall_gap.quantize(Decimal("0.01")),
            "diagnostics": result.diagnostics,
            "teams": team_payload,
        }
        response_serializer = TeamGenerationResponseSerializer(payload)
        return Response(response_serializer.data)

    @action(detail=True, methods=["post"], url_path="clear-teams")
    def clear_teams(self, request, pk=None):
        match = self.get_object()

        with transaction.atomic():
            match.attendance_entries.update(
                assigned_team_number=None,
                assigned_team_name="",
            )
            match.teams_generated_at = None
            match.save(update_fields=["teams_generated_at", "updated_at"])

        return Response(self.get_serializer(match).data)

    @action(detail=True, methods=["post"], url_path="swap-team-players")
    def swap_team_players(self, request, pk=None):
        serializer = TeamPlayerSwapSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        match = self.get_object()
        source_id = serializer.validated_data["source_attendance_id"]
        target_id = serializer.validated_data["target_attendance_id"]

        with transaction.atomic():
            entries = {
                entry.id: entry
                for entry in MatchAttendance.objects.select_for_update().filter(
                    match=match,
                    id__in=[source_id, target_id],
                    attendance_status=MatchAttendance.AttendanceStatus.CONFIRMED,
                )
            }
            source_entry = entries.get(source_id)
            target_entry = entries.get(target_id)

            if source_entry is None or target_entry is None:
                raise ValidationError({"detail": "Selecione jogadores confirmados desta pelada."})

            if (
                source_entry.assigned_team_number is None
                or target_entry.assigned_team_number is None
            ):
                raise ValidationError({"detail": "Gere os times antes de trocar jogadores."})

            if source_entry.assigned_team_number == target_entry.assigned_team_number:
                raise ValidationError({"detail": "Selecione jogadores de times diferentes."})

            source_team_number = source_entry.assigned_team_number
            source_team_name = source_entry.assigned_team_name
            source_entry.assigned_team_number = target_entry.assigned_team_number
            source_entry.assigned_team_name = target_entry.assigned_team_name
            target_entry.assigned_team_number = source_team_number
            target_entry.assigned_team_name = source_team_name

            source_entry.save(
                update_fields=["assigned_team_number", "assigned_team_name", "updated_at"]
            )
            target_entry.save(
                update_fields=["assigned_team_number", "assigned_team_name", "updated_at"]
            )
            match.updated_at = timezone.now()
            match.save(update_fields=["updated_at"])

        refreshed_attendance = match.attendance_entries.select_related(
            "match", "player", "invited_by"
        ).order_by(
            "display_name",
        )
        return Response(MatchAttendanceSerializer(refreshed_attendance, many=True).data)

    @action(detail=True, methods=["post"], url_path="finalize")
    def finalize(self, request, pk=None):
        input_serializer = MatchFinalizeSerializer(data=request.data)
        input_serializer.is_valid(raise_exception=True)
        match = self.get_object()
        finalize_match(
            match,
            input_serializer.validated_data.get("winning_team_number"),
            request.user,
        )
        return Response(self.get_serializer(match).data)

    @action(detail=True, methods=["get"], url_path="stats-sheet")
    def export_stats_sheet(self, request, pk=None):
        match = self.get_object()
        return build_match_stats_xlsx_response(match)

    @action(
        detail=True,
        methods=["post"],
        url_path="import-stats-sheet",
        parser_classes=[MultiPartParser, FormParser],
    )
    def import_stats_sheet(self, request, pk=None):
        match = self.get_object()
        uploaded_file = request.FILES.get("file")
        if uploaded_file is None:
            raise ValidationError({"file": "Envie um arquivo XLSX no campo file."})

        summary = import_match_stats_from_xlsx(match, uploaded_file, request.user)
        serializer = MatchStatsImportSummarySerializer(summary)
        return Response(serializer.data)

    @action(detail=False, methods=["get"], url_path="overall-history")
    def overall_history(self, request):
        finalize_due_match_ratings()
        member_players = Player.objects.filter(player_type=Player.PlayerType.MEMBER).order_by(
            "full_name"
        )
        players_payload = [
            {
                "player_id": str(player.id),
                "display_name": player.nickname or player.full_name,
                "is_active": player.is_active,
            }
            for player in member_players
        ]
        member_player_ids = [player.id for player in member_players]
        entries = list(
            MatchAttendance.objects.select_related("match", "player")
            .filter(
                match__status__in=FINAL_MATCH_STATUSES,
                match__scheduled_at__date__gte=OVERALL_HISTORY_START_DATE,
                attendance_status=MatchAttendance.AttendanceStatus.CONFIRMED,
                player_id__in=member_player_ids,
            )
            .exclude(match__scheduled_at__date__in=OVERALL_HISTORY_EXCLUDED_DATES)
            .order_by("match__scheduled_at", "display_name")
        )
        entries_by_player_id = {}
        for entry in entries:
            entries_by_player_id.setdefault(entry.player_id, []).append(entry)

        overall_after_match_by_attendance_id = {}
        for player_entries in entries_by_player_id.values():
            for index, entry in enumerate(player_entries):
                next_entry = player_entries[index + 1] if index + 1 < len(player_entries) else None
                overall_after_match_by_attendance_id[entry.id] = (
                    next_entry.overall if next_entry else entry.player.overall
                )

        matches_by_id = {}
        for entry in entries:
            match_payload = matches_by_id.setdefault(
                entry.match_id,
                {
                    "match_id": str(entry.match_id),
                    "scheduled_at": entry.match.scheduled_at,
                    "location": entry.match.location,
                    "points": [],
                },
            )
            match_payload["points"].append(
                {
                    "player_id": str(entry.player_id),
                    "display_name": entry.player.nickname or entry.player.full_name,
                    "overall": overall_after_match_by_attendance_id[entry.id],
                }
            )

        payload = {
            "players": players_payload,
            "matches": list(matches_by_id.values()),
        }
        return Response(payload)

    def _get_rating_lock_reason(self, match):
        if match.status != Match.Status.ARCHIVED:
            return "Avaliações liberadas após o arquivamento da pelada."

        if match.ratings_finalized_at:
            return "A janela de notas desta pelada foi encerrada."

        if is_rating_window_expired(match):
            return "A janela de notas desta pelada foi encerrada."

        return ""

    def _build_rating_overall_summary(self, match, rating_stats):
        summary = []
        participants = (
            match.attendance_entries.select_related("player")
            .filter(
                attendance_status=MatchAttendance.AttendanceStatus.CONFIRMED,
                player__isnull=False,
            )
            .order_by("display_name")
        )
        for entry in participants:
            stats = rating_stats.get(entry.id, {})
            current_overall = entry.player.overall
            summary.append(
                {
                    "attendance_id": entry.id,
                    "player_id": entry.player_id,
                    "display_name": entry.display_name,
                    "previous_overall": entry.overall,
                    "current_overall": current_overall,
                    "delta": current_overall - entry.overall,
                    "average_score": stats.get("average_score"),
                    "rating_count": stats.get("rating_count", 0),
                }
            )
        return summary

    def _build_rating_log(self, match):
        rating_log = []
        for rating in (
            MatchPlayerRating.objects.filter(match=match)
            .select_related("rater_user", "rater", "rated_attendance", "rated_player")
            .order_by("-created_at")
        ):
            rater_display_name = ""
            if rating.rater_user_id:
                rater_display_name = (
                    rating.rater_user.display_name
                    or rating.rater_user.get_full_name()
                    or rating.rater_user.username
                )
            elif rating.rater_id:
                rater_display_name = rating.rater.full_name

            rating_log.append(
                {
                    "rater_user_id": rating.rater_user_id,
                    "rater_display_name": rater_display_name or "Usuário sem nome",
                    "rated_attendance_id": rating.rated_attendance_id,
                    "rated_player_id": rating.rated_player_id,
                    "rated_display_name": rating.rated_attendance.display_name,
                    "score": rating.score,
                    "created_at": rating.created_at,
                    "updated_at": rating.updated_at,
                }
            )
        return rating_log

    def _get_rating_participants_queryset(self, match):
        return match.attendance_entries.select_related("player").filter(
            attendance_status=MatchAttendance.AttendanceStatus.CONFIRMED,
            player__isnull=False,
        )

    def _get_rateable_entries_queryset(self, match, linked_player):
        """Entries the given player may rate, according to the match rating mode.

        In GENERAL mode everyone rates everyone except themselves; in TEAM mode
        the ballot is restricted to the rater's own team.
        """
        participants_queryset = self._get_rating_participants_queryset(match)
        if match.rating_mode == Match.RatingMode.GENERAL:
            return participants_queryset.exclude(player=linked_player), ""

        linked_attendance = participants_queryset.filter(player=linked_player).first()
        if not linked_attendance or linked_attendance.assigned_team_number is None:
            return participants_queryset.none(), MISSING_TEAM_RATING_LOCK_REASON

        return (
            participants_queryset.filter(
                assigned_team_number=linked_attendance.assigned_team_number,
            ).exclude(player=linked_player),
            "",
        )

    def _build_rating_state(self, match):
        finalize_match_ratings_if_due(match)
        match.refresh_from_db(fields=["ratings_finalized_at"])
        window_closes_at = get_rating_window_closes_at(match)
        rating_stats = get_match_rating_stats(match)
        overall_summary = self._build_rating_overall_summary(match, rating_stats)
        rating_log = self._build_rating_log(match)
        locked_reason = self._get_rating_lock_reason(match)
        linked_player = getattr(self.request.user, "linked_player", None)
        if (
            getattr(self.request.user, "role", None) != Role.ADMIN
            and not locked_reason
            and not linked_player
        ):
            locked_reason = UNLINKED_RATING_LOCK_REASON
        can_rate = bool(
            getattr(self.request.user, "role", None) != Role.ADMIN and not locked_reason
        )
        is_window_closed = bool(
            match.status == Match.Status.ARCHIVED
            and (match.ratings_finalized_at or is_rating_window_expired(match))
        )

        if is_window_closed:
            payload = {
                "match_id": match.id,
                "can_rate": False,
                "has_submitted": False,
                "locked_reason": locked_reason,
                "window_closes_at": window_closes_at,
                "ratings_finalized_at": match.ratings_finalized_at,
                "items": [],
                "log": rating_log,
                "overall_summary": overall_summary,
            }
            return MatchPlayerRatingStateSerializer(payload).data

        participants_queryset = self._get_rating_participants_queryset(match)
        if linked_player:
            participants_queryset, ballot_locked_reason = self._get_rateable_entries_queryset(
                match,
                linked_player,
            )
            if can_rate and ballot_locked_reason:
                locked_reason = ballot_locked_reason
                can_rate = False

        participants = list(participants_queryset.order_by("display_name"))
        participant_ids = [entry.id for entry in participants]
        rating_by_attendance = {
            rating.rated_attendance_id: rating
            for rating in MatchPlayerRating.objects.filter(
                match=match,
                rater_user=self.request.user,
                rated_attendance_id__in=participant_ids,
            )
        }

        items = []
        for entry in participants:
            existing_rating = rating_by_attendance.get(entry.id)
            stats = rating_stats.get(entry.id, {})
            items.append(
                {
                    "attendance_id": entry.id,
                    "player_id": entry.player_id,
                    "display_name": entry.display_name,
                    "current_overall": entry.player.overall,
                    "score": existing_rating.score if existing_rating else None,
                    "average_score": stats.get("average_score"),
                    "rating_count": stats.get("rating_count", 0),
                }
            )

        if can_rate and not items:
            locked_reason = (
                "Nao ha outros participantes confirmados para avaliar."
                if match.rating_mode == Match.RatingMode.GENERAL
                else "Nao ha outros mensalistas confirmados para avaliar."
            )
            can_rate = False

        payload = {
            "match_id": match.id,
            "can_rate": can_rate,
            "has_submitted": bool(rating_by_attendance),
            "locked_reason": locked_reason,
            "window_closes_at": window_closes_at,
            "ratings_finalized_at": match.ratings_finalized_at,
            "items": items,
            "log": rating_log,
            "overall_summary": overall_summary,
        }
        return MatchPlayerRatingStateSerializer(payload).data

    @action(detail=True, methods=["get", "post"], url_path="player-ratings")
    def player_ratings(self, request, pk=None):
        match = self.get_object()
        if getattr(request.user, "role", None) == Role.ADMIN:
            return Response(self._build_rating_state(match))

        locked_reason = self._get_rating_lock_reason(match)
        if request.method == "GET":
            return Response(self._build_rating_state(match))

        if locked_reason:
            raise ValidationError({"detail": locked_reason or "Avaliação indisponível."})

        linked_player = getattr(request.user, "linked_player", None)
        if not linked_player:
            raise ValidationError({"detail": UNLINKED_RATING_LOCK_REASON})

        serializer = MatchPlayerRatingSubmitSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        allowed_entries_queryset, ballot_locked_reason = self._get_rateable_entries_queryset(
            match,
            linked_player,
        )
        if ballot_locked_reason:
            raise ValidationError({"detail": ballot_locked_reason})

        invalid_rating_message = (
            "Avalie apenas outros participantes confirmados desta pelada."
            if match.rating_mode == Match.RatingMode.GENERAL
            else "Avalie apenas jogadores do mesmo time do seu jogador vinculado."
        )
        allowed_entries = {entry.id: entry for entry in allowed_entries_queryset}
        with transaction.atomic():
            for item in serializer.validated_data["ratings"]:
                attendance_id = item["attendance_id"]
                attendance_entry = allowed_entries.get(attendance_id)
                if not attendance_entry:
                    raise ValidationError({"ratings": invalid_rating_message})
                if item["score"] is None:
                    # Voto pulado: o votante optou por nao avaliar este jogador.
                    MatchPlayerRating.objects.filter(
                        match=match,
                        rater_user=request.user,
                        rated_attendance=attendance_entry,
                    ).delete()
                    continue
                MatchPlayerRating.objects.update_or_create(
                    match=match,
                    rater_user=request.user,
                    rated_attendance=attendance_entry,
                    defaults={
                        "rater": linked_player,
                        "rated_player": attendance_entry.player,
                        "score": item["score"],
                    },
                )

        return Response(self._build_rating_state(match))

    @action(detail=True, methods=["post"], url_path="finalize-ratings")
    def finalize_ratings(self, request, pk=None):
        match = self.get_object()
        if match.status != Match.Status.ARCHIVED:
            raise ValidationError({"detail": "Arquive a pelada antes de finalizar as notas."})

        finalize_match_ratings(match, force=True)
        return Response(self._build_rating_state(match))

    @action(detail=True, methods=["post"], url_path="recalculate-ratings")
    def recalculate_ratings(self, request, pk=None):
        match = self.get_object()
        if match.status != Match.Status.ARCHIVED:
            raise ValidationError({"detail": "Arquive a pelada antes de recalcular as notas."})
        if not match.ratings_finalized_at:
            raise ValidationError(
                {"detail": "Finalize a janela de notas antes de recalcular os overalls."}
            )

        recalculate_finalized_match_ratings(match)
        return Response(self._build_rating_state(match))


class TransactionViewSet(viewsets.ModelViewSet):
    queryset = Transaction.objects.select_related(
        "related_player",
        "match",
        "recorded_by",
    ).all()
    serializer_class = TransactionSerializer
    permission_classes = [IsAdminOrReadOnly]

    def get_queryset(self):
        queryset = super().get_queryset()
        if getattr(self.request.user, "role", None) == Role.ADMIN:
            return queryset

        linked_player = getattr(self.request.user, "linked_player", None)
        if not linked_player:
            return queryset.none()
        return queryset.filter(related_player=linked_player)

    def perform_create(self, serializer):
        serializer.save(recorded_by=self.request.user)


class MatchAttendanceViewSet(viewsets.ModelViewSet):
    queryset = MatchAttendance.objects.select_related(
        "match",
        "player",
        "invited_by",
    ).all()
    serializer_class = MatchAttendanceSerializer
    permission_classes = [IsAdminOrReadOnly]

    def get_queryset(self):
        queryset = super().get_queryset()
        if getattr(self.request.user, "role", None) != Role.ADMIN:
            queryset = queryset.filter(
                match__status__in=[Match.Status.OPEN, Match.Status.ARCHIVED]
            ).distinct()
        guest_fee_due = self.request.query_params.get("guest_fee_due")
        if guest_fee_due in {"1", "true", "True"}:
            return queryset.filter(
                is_guest=True,
                attendance_status=MatchAttendance.AttendanceStatus.CONFIRMED,
                guest_fee_status=MatchAttendance.GuestFeeStatus.PENDING,
                guest_fee_amount__gt=Decimal("0.00"),
                match__status__in=FINAL_MATCH_STATUSES,
            ).order_by("-match__scheduled_at", "display_name")
        match_id = self.request.query_params.get("match")
        if match_id:
            queryset = queryset.filter(match_id=match_id)
        return queryset.order_by("display_name")

    def _build_attendance_payload(self, validated_data, instance=None):
        player = validated_data.get("player")
        attendance_status = validated_data.get(
            "attendance_status",
            instance.attendance_status if instance else MatchAttendance.AttendanceStatus.CONFIRMED,
        )

        if player and not validated_data.get("display_name"):
            validated_data["display_name"] = player.full_name

        if player:
            validated_data.setdefault("overall", player.overall)

        is_guest = validated_data.get("is_guest", instance.is_guest if instance else False)
        if is_guest:
            validated_data.setdefault("guest_fee_amount", GUEST_FEE_AMOUNT)
            validated_data.setdefault("guest_fee_status", MatchAttendance.GuestFeeStatus.PENDING)
        else:
            validated_data["guest_fee_amount"] = Decimal("0.00")
            validated_data["guest_fee_status"] = MatchAttendance.GuestFeeStatus.WAIVED
            validated_data["guest_fee_paid_at"] = None

        if "attendance_status" in validated_data or instance is None:
            validated_data["confirmed_at"] = (
                timezone.now()
                if attendance_status == MatchAttendance.AttendanceStatus.CONFIRMED
                else None
            )
        return validated_data

    def perform_create(self, serializer):
        serializer.save(**self._build_attendance_payload(dict(serializer.validated_data)))

    def perform_update(self, serializer):
        serializer.save(
            **self._build_attendance_payload(
                dict(serializer.validated_data),
                instance=serializer.instance,
            )
        )

    @action(detail=True, methods=["post"], url_path="mark-guest-fee-paid")
    def mark_guest_fee_paid(self, request, pk=None):
        attendance = self.get_object()
        if not attendance.is_guest:
            raise ValidationError({"detail": "Apenas convidados possuem taxa avulsa."})
        if attendance.attendance_status != MatchAttendance.AttendanceStatus.CONFIRMED:
            raise ValidationError({"detail": "A taxa só é cobrada de convidados confirmados."})
        if attendance.match.status not in FINAL_MATCH_STATUSES:
            raise ValidationError(
                {"detail": "A taxa do convidado só é finalizada ao fim da pelada."}
            )
        if attendance.guest_fee_status == MatchAttendance.GuestFeeStatus.PAID:
            return Response(self.get_serializer(attendance).data)

        with transaction.atomic():
            attendance = (
                MatchAttendance.objects.select_for_update()
                .select_related("match")
                .get(pk=attendance.pk)
            )
            if attendance.guest_fee_status != MatchAttendance.GuestFeeStatus.PAID:
                Transaction.objects.create(
                    direction=Transaction.Direction.INFLOW,
                    category=Transaction.Category.EXTRA_FEE,
                    status=Transaction.Status.POSTED,
                    amount=attendance.guest_fee_amount,
                    description=f"Taxa de convidado - {attendance.display_name}",
                    occurred_on=timezone.localdate(),
                    reference_month=timezone.localdate().replace(day=1),
                    related_player=attendance.invited_by,
                    match=attendance.match,
                    recorded_by=request.user,
                    external_reference=f"guest-fee:{attendance.id}",
                    notes=f"Taxa avulsa do convidado {attendance.display_name}.",
                )
                attendance.guest_fee_status = MatchAttendance.GuestFeeStatus.PAID
                attendance.guest_fee_paid_at = timezone.now()
                attendance.save(
                    update_fields=["guest_fee_status", "guest_fee_paid_at", "updated_at"]
                )

        return Response(self.get_serializer(attendance).data)

    @action(detail=True, methods=["post"], url_path="waive-guest-fee")
    def waive_guest_fee(self, request, pk=None):
        attendance = self.get_object()
        if not attendance.is_guest:
            raise ValidationError({"detail": "Apenas convidados possuem taxa avulsa."})
        if attendance.attendance_status != MatchAttendance.AttendanceStatus.CONFIRMED:
            raise ValidationError({"detail": "A taxa só é cobrada de convidados confirmados."})
        if attendance.match.status not in FINAL_MATCH_STATUSES:
            raise ValidationError(
                {"detail": "A taxa do convidado só é finalizada ao fim da pelada."}
            )
        if attendance.guest_fee_status == MatchAttendance.GuestFeeStatus.PAID:
            raise ValidationError({"detail": "A taxa do convidado ja foi marcada como paga."})
        if attendance.guest_fee_status == MatchAttendance.GuestFeeStatus.WAIVED:
            return Response(self.get_serializer(attendance).data)

        with transaction.atomic():
            attendance = (
                MatchAttendance.objects.select_for_update()
                .select_related("match")
                .get(pk=attendance.pk)
            )
            if attendance.guest_fee_status == MatchAttendance.GuestFeeStatus.PENDING:
                attendance.guest_fee_status = MatchAttendance.GuestFeeStatus.WAIVED
                attendance.guest_fee_paid_at = None
                attendance.save(
                    update_fields=["guest_fee_status", "guest_fee_paid_at", "updated_at"]
                )

        return Response(self.get_serializer(attendance).data)


class AuthLoginView(APIView):
    permission_classes = [AllowAny]
    authentication_classes = []

    def post(self, request):
        serializer = LoginSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.validated_data["user"]
        token, _ = Token.objects.get_or_create(user=user)
        return Response({"token": token.key, "user": UserSerializer(user).data})


class AuthSignupView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = PublicSignupSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        token = Token.objects.create(user=user)
        return Response(
            {"token": token.key, "user": UserSerializer(user).data},
            status=status.HTTP_201_CREATED,
        )


class AuthMeView(APIView):
    def get(self, request):
        return Response(UserSerializer(request.user).data)

    def patch(self, request):
        serializer = UserProfileUpdateSerializer(
            request.user,
            data=request.data,
            partial=True,
            context={"request": request},
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(UserSerializer(request.user).data)


class AuthLogoutView(APIView):
    def post(self, request):
        Token.objects.filter(user=request.user).delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class AuthChangePasswordView(APIView):
    def post(self, request):
        serializer = ChangePasswordSerializer(data=request.data, context={"request": request})
        serializer.is_valid(raise_exception=True)

        user = request.user
        user.set_password(serializer.validated_data["new_password"])
        user.must_change_password = False
        user.save()

        Token.objects.filter(user=user).delete()
        token = Token.objects.create(user=user)
        return Response({"token": token.key, "user": UserSerializer(user).data})


def build_cash_totals():
    """Totais do caixa do grupo, usados pelo dashboard e pelo painel do jogador."""
    totals = Transaction.objects.aggregate(
        inflow_total=Coalesce(
            Sum(
                Case(
                    When(
                        direction=Transaction.Direction.INFLOW,
                        status=Transaction.Status.POSTED,
                        then=F("amount"),
                    ),
                    default=Value(0),
                    output_field=DecimalField(max_digits=10, decimal_places=2),
                )
            ),
            Decimal("0.00"),
        ),
        outflow_total=Coalesce(
            Sum(
                Case(
                    When(
                        direction=Transaction.Direction.OUTFLOW,
                        status=Transaction.Status.POSTED,
                        then=F("amount"),
                    ),
                    default=Value(0),
                    output_field=DecimalField(max_digits=10, decimal_places=2),
                )
            ),
            Decimal("0.00"),
        ),
        pending_total=Coalesce(
            Sum(
                Case(
                    When(status=Transaction.Status.PENDING, then=F("amount")),
                    default=Value(0),
                    output_field=DecimalField(max_digits=10, decimal_places=2),
                )
            ),
            Decimal("0.00"),
        ),
    )
    totals["pending_total"] += get_guest_fee_pending_total()
    totals["current_balance"] = totals["inflow_total"] - totals["outflow_total"]
    return totals


class FinancialSummaryView(APIView):
    permission_classes = [IsRoleAdmin]

    def get(self, request):
        serializer = FinancialSummarySerializer(build_cash_totals())
        return Response(serializer.data)


class PortalOverviewView(APIView):
    def get(self, request):
        user = request.user
        linked_player = user.linked_player
        month_start, month_end = month_bounds(timezone.localdate())
        reference_month = month_start.strftime("%Y-%m")

        financial_status = {
            "reference_month": reference_month,
            "expected_monthly_fee": Decimal("0.00"),
            "paid_amount": Decimal("0.00"),
            "pending_amount": Decimal("0.00"),
            "outstanding_amount": Decimal("0.00"),
            "is_adimplente": False,
        }
        attendance_status = {
            "confirmed_count": 0,
            "pending_count": 0,
            "declined_count": 0,
            "total_count": 0,
        }
        recent_attendance = []

        if linked_player:
            monthly_map = build_player_payment_map([linked_player.id], month_start, month_end)
            paid_amount = monthly_map.get(linked_player.id, {}).get("paid_amount", Decimal("0.00"))
            pending_amount = monthly_map.get(linked_player.id, {}).get(
                "pending_amount", Decimal("0.00")
            )
            expected_monthly_fee = linked_player.monthly_fee_amount
            outstanding_amount = expected_monthly_fee - paid_amount
            if outstanding_amount < Decimal("0.00"):
                outstanding_amount = Decimal("0.00")

            financial_status = {
                "reference_month": reference_month,
                "expected_monthly_fee": expected_monthly_fee,
                "paid_amount": paid_amount,
                "pending_amount": pending_amount,
                "outstanding_amount": outstanding_amount,
                "is_adimplente": outstanding_amount == Decimal("0.00"),
            }

            attendance_counts = MatchAttendance.objects.filter(player=linked_player).aggregate(
                confirmed_count=Count(
                    "id",
                    filter=Q(attendance_status=MatchAttendance.AttendanceStatus.CONFIRMED),
                ),
                pending_count=Count(
                    "id",
                    filter=Q(attendance_status=MatchAttendance.AttendanceStatus.PENDING),
                ),
                declined_count=Count(
                    "id",
                    filter=Q(attendance_status=MatchAttendance.AttendanceStatus.DECLINED),
                ),
                total_count=Count("id"),
            )
            attendance_status = {
                "confirmed_count": attendance_counts["confirmed_count"],
                "pending_count": attendance_counts["pending_count"],
                "declined_count": attendance_counts["declined_count"],
                "total_count": attendance_counts["total_count"],
            }

            attendance_entries = (
                MatchAttendance.objects.select_related("match")
                .filter(player=linked_player)
                .order_by("-match__scheduled_at")[:5]
            )
            recent_attendance = [
                {
                    "match_id": entry.match_id,
                    "scheduled_at": entry.match.scheduled_at,
                    "match_status": entry.match.status,
                    "attendance_status": entry.attendance_status,
                    "assigned_team_name": entry.assigned_team_name,
                }
                for entry in attendance_entries
            ]

        upcoming_matches_qs = Match.objects.filter(
            status=Match.Status.OPEN, scheduled_at__gte=timezone.now()
        ).order_by("scheduled_at")[:5]
        upcoming_matches = list(upcoming_matches_qs)
        attendance_by_match = {}
        if linked_player and upcoming_matches:
            attendance_by_match = {
                row["match_id"]: row["attendance_status"]
                for row in MatchAttendance.objects.filter(
                    player=linked_player,
                    match_id__in=[match.id for match in upcoming_matches],
                ).values("match_id", "attendance_status")
            }

        upcoming_payload = [
            {
                "match_id": match.id,
                "scheduled_at": match.scheduled_at,
                "location": match.location,
                "status": match.status,
                "expected_team_count": match.expected_team_count,
                "attendance_status": attendance_by_match.get(match.id),
            }
            for match in upcoming_matches
        ]

        cash_totals = build_cash_totals()

        payload = {
            "user": user,
            "linked_player": linked_player,
            "financial_status": financial_status,
            "attendance_status": attendance_status,
            "recent_attendance": recent_attendance,
            "upcoming_matches": upcoming_payload,
            "cash": {
                "current_balance": cash_totals["current_balance"],
                "pending_total": cash_totals["pending_total"],
            },
        }
        serializer = PortalOverviewSerializer(payload)
        return Response(serializer.data)


class SeasonOverviewView(APIView):
    permission_classes = [IsRoleAdmin]

    def get(self, request):
        month_start, month_end = month_bounds(timezone.localdate())
        reference_month = month_start.strftime("%Y-%m")

        match_counts = Match.objects.aggregate(
            total_matches=Count("id"),
            matches_open=Count("id", filter=Q(status=Match.Status.OPEN)),
            matches_archived=Count("id", filter=Q(status=Match.Status.ARCHIVED)),
        )

        attendance_counts = MatchAttendance.objects.aggregate(
            attendance_confirmed=Count(
                "id",
                filter=Q(attendance_status=MatchAttendance.AttendanceStatus.CONFIRMED),
            ),
            attendance_pending=Count(
                "id",
                filter=Q(attendance_status=MatchAttendance.AttendanceStatus.PENDING),
            ),
            attendance_declined=Count(
                "id",
                filter=Q(attendance_status=MatchAttendance.AttendanceStatus.DECLINED),
            ),
            attendance_total=Count("id"),
        )

        finance_totals = Transaction.objects.aggregate(
            inflow_total=Coalesce(
                Sum(
                    Case(
                        When(
                            direction=Transaction.Direction.INFLOW,
                            status=Transaction.Status.POSTED,
                            then=F("amount"),
                        ),
                        default=Value(0),
                        output_field=DecimalField(max_digits=12, decimal_places=2),
                    )
                ),
                Decimal("0.00"),
            ),
            outflow_total=Coalesce(
                Sum(
                    Case(
                        When(
                            direction=Transaction.Direction.OUTFLOW,
                            status=Transaction.Status.POSTED,
                            then=F("amount"),
                        ),
                        default=Value(0),
                        output_field=DecimalField(max_digits=12, decimal_places=2),
                    )
                ),
                Decimal("0.00"),
            ),
            pending_total=Coalesce(
                Sum(
                    Case(
                        When(status=Transaction.Status.PENDING, then=F("amount")),
                        default=Value(0),
                        output_field=DecimalField(max_digits=12, decimal_places=2),
                    )
                ),
                Decimal("0.00"),
            ),
        )
        finance_totals["pending_total"] += get_guest_fee_pending_total()

        active_members = list(
            Player.objects.filter(
                player_type=Player.PlayerType.MEMBER,
                is_active=True,
            ).only("id", "monthly_fee_amount")
        )
        payment_map = build_player_payment_map(
            [player.id for player in active_members],
            month_start,
            month_end,
        )

        adimplent_members = 0
        for player in active_members:
            paid_amount = payment_map.get(player.id, {}).get("paid_amount", Decimal("0.00"))
            outstanding = player.monthly_fee_amount - paid_amount
            if outstanding <= Decimal("0.00"):
                adimplent_members += 1

        payload = {
            "reference_month": reference_month,
            **match_counts,
            "active_members": len(active_members),
            **attendance_counts,
            **finance_totals,
            "current_balance": finance_totals["inflow_total"] - finance_totals["outflow_total"],
            "adimplent_members": adimplent_members,
            "delinquent_members": len(active_members) - adimplent_members,
        }
        serializer = SeasonOverviewSerializer(payload)
        return Response(serializer.data)


class PresenceRankingView(APIView):
    permission_classes = [IsRoleAdmin]

    def get(self, request):
        try:
            limit = int(request.query_params.get("limit", "20"))
        except ValueError as exc:
            raise ValidationError({"limit": "Forneca um numero inteiro."}) from exc

        ranked_players = (
            Player.objects.filter(player_type=Player.PlayerType.MEMBER, is_active=True)
            .annotate(
                confirmed_count=Count(
                    "attendance_entries",
                    filter=Q(
                        attendance_entries__attendance_status=MatchAttendance.AttendanceStatus.CONFIRMED
                    ),
                ),
                pending_count=Count(
                    "attendance_entries",
                    filter=Q(
                        attendance_entries__attendance_status=MatchAttendance.AttendanceStatus.PENDING
                    ),
                ),
                declined_count=Count(
                    "attendance_entries",
                    filter=Q(
                        attendance_entries__attendance_status=MatchAttendance.AttendanceStatus.DECLINED
                    ),
                ),
                total_calls=Count("attendance_entries"),
            )
            .order_by("-confirmed_count", "full_name")[: max(limit, 1)]
        )

        ranking = []
        for player in ranked_players:
            attendance_rate = Decimal("0.00")
            if player.total_calls:
                attendance_rate = (
                    Decimal(player.confirmed_count)
                    * Decimal("100.00")
                    / Decimal(player.total_calls)
                ).quantize(Decimal("0.01"))

            ranking.append(
                {
                    "player_id": player.id,
                    "player_name": player.full_name,
                    "confirmed_count": player.confirmed_count,
                    "pending_count": player.pending_count,
                    "declined_count": player.declined_count,
                    "total_calls": player.total_calls,
                    "attendance_rate": attendance_rate,
                }
            )

        serializer = PresenceRankingSerializer({"ranking": ranking})
        return Response(serializer.data)


class PaymentRankingView(APIView):
    permission_classes = [IsRoleAdmin]

    def get(self, request):
        month_start, reference_month = parse_reference_month(
            request.query_params.get("reference_month")
        )
        _, month_end = month_bounds(month_start)

        try:
            limit = int(request.query_params.get("limit", "20"))
        except ValueError as exc:
            raise ValidationError({"limit": "Forneca um numero inteiro."}) from exc

        members = list(
            Player.objects.filter(player_type=Player.PlayerType.MEMBER, is_active=True).only(
                "id",
                "full_name",
                "monthly_fee_amount",
            )
        )
        payment_map = build_player_payment_map(
            [player.id for player in members],
            month_start,
            month_end,
        )

        ranking = []
        for player in members:
            paid_amount = payment_map.get(player.id, {}).get("paid_amount", Decimal("0.00"))
            pending_amount = payment_map.get(player.id, {}).get("pending_amount", Decimal("0.00"))
            outstanding_amount = player.monthly_fee_amount - paid_amount
            if outstanding_amount < Decimal("0.00"):
                outstanding_amount = Decimal("0.00")

            ranking.append(
                {
                    "player_id": player.id,
                    "player_name": player.full_name,
                    "expected_monthly_fee": player.monthly_fee_amount,
                    "paid_amount": paid_amount,
                    "pending_amount": pending_amount,
                    "outstanding_amount": outstanding_amount,
                    "is_adimplente": outstanding_amount == Decimal("0.00"),
                }
            )

        ranking.sort(
            key=lambda item: (
                item["outstanding_amount"],
                item["pending_amount"],
                item["player_name"],
            )
        )
        payload = {
            "reference_month": reference_month,
            "ranking": ranking[: max(limit, 1)],
        }
        serializer = PaymentRankingSerializer(payload)
        return Response(serializer.data)


class SportsRankingView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            limit = int(request.query_params.get("limit", "20"))
        except ValueError as exc:
            raise ValidationError({"limit": "Forneca um numero inteiro."}) from exc

        payload = build_sports_ranking(max(1, min(limit, 50)))
        serializer = SportsRankingSerializer(payload)
        return Response(serializer.data)
