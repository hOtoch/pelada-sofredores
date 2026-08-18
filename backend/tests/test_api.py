from __future__ import annotations

import io
from datetime import date, datetime, timedelta
from decimal import Decimal

from django.core.files.uploadedfile import SimpleUploadedFile
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework.authtoken.models import Token
from rest_framework.test import APITestCase

from core.models import Match, MatchAttendance, MatchPlayerRating, MatchPlayerStat, Player, Role, Transaction, User


class ApiFlowTests(APITestCase):
    def setUp(self) -> None:
        self.password = "pelada-segura-123"
        self.user = User.objects.create_user(
            username="admin",
            email="admin@pelada.local",
            password=self.password,
            role=Role.ADMIN,
            display_name="Administrador",
        )
        self.token = Token.objects.create(user=self.user)
        self.client.credentials(HTTP_AUTHORIZATION=f"Token {self.token.key}")
        self.common_user = User.objects.create_user(
            username="comum",
            email="comum@pelada.local",
            password=self.password,
            role=Role.COMMON,
            display_name="Jogador Comum",
        )
        self.common_token = Token.objects.create(user=self.common_user)

        self.match = Match.objects.create(
            scheduled_at=timezone.now() + timedelta(days=2),
            status=Match.Status.OPEN,
            expected_team_count=2,
            created_by=self.user,
        )

        self.players = [
            Player.objects.create(
                full_name=f"Jogador {index}",
                preferred_position=Player.PreferredPosition.UNIVERSAL,
                overall=80 - index,
            )
            for index in range(4)
        ]

        for player in self.players:
            MatchAttendance.objects.create(
                match=self.match,
                player=player,
                display_name=player.full_name,
                is_guest=False,
                attendance_status=MatchAttendance.AttendanceStatus.CONFIRMED,
                overall=player.overall,
            )

        Transaction.objects.create(
            direction=Transaction.Direction.INFLOW,
            category=Transaction.Category.MONTHLY_FEE,
            status=Transaction.Status.POSTED,
            amount=Decimal("300.00"),
            description="Mensalidade",
            occurred_on=date.today(),
            recorded_by=self.user,
        )
        Transaction.objects.create(
            direction=Transaction.Direction.OUTFLOW,
            category=Transaction.Category.FIELD_RENT,
            status=Transaction.Status.POSTED,
            amount=Decimal("120.00"),
            description="Campo",
            occurred_on=date.today(),
            recorded_by=self.user,
        )
        Transaction.objects.create(
            direction=Transaction.Direction.OUTFLOW,
            category=Transaction.Category.BARBECUE,
            status=Transaction.Status.PENDING,
            amount=Decimal("45.00"),
            description="Churrasco",
            occurred_on=date.today(),
            recorded_by=self.user,
        )

    def test_login_endpoint_returns_token_and_user(self) -> None:
        self.client.credentials()

        response = self.client.post(
            "/api/auth/login/",
            {"identifier": self.user.email, "password": self.password},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn("token", response.data)
        self.assertEqual(response.data["user"]["username"], self.user.username)

    def test_login_with_phone_number(self) -> None:
        self.client.credentials()
        self.common_user.phone_number = "27999998888"
        self.common_user.save(update_fields=["phone_number"])

        response = self.client.post(
            "/api/auth/login/",
            {"identifier": "(27) 99999-8888", "password": self.password},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["user"]["username"], self.common_user.username)

    def test_public_signup_creates_common_user_without_linked_player(self) -> None:
        self.client.credentials()

        response = self.client.post(
            "/api/auth/register/",
            {
                "full_name": "Novo Jogador",
                "phone_number": "(27) 98888-7777",
                "username": "novo_jogador",
                "password": "SenhaNova123",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201)
        self.assertIn("token", response.data)
        created_user = User.objects.get(username="novo_jogador")
        self.assertEqual(created_user.role, Role.COMMON)
        self.assertEqual(created_user.phone_number, "27988887777")
        self.assertIsNone(created_user.linked_player)
        self.assertFalse(Player.objects.filter(full_name="Novo Jogador").exists())

    def test_public_signup_rejects_username_with_spaces(self) -> None:
        self.client.credentials()

        response = self.client.post(
            "/api/auth/register/",
            {
                "full_name": "Novo Jogador",
                "phone_number": "(27) 98888-7777",
                "username": "novo jogador",
                "password": "SenhaNova123",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("username", response.data)

    def test_public_signup_allows_phone_from_existing_player(self) -> None:
        self.client.credentials()
        self.players[0].phone_number = "27988887777"
        self.players[0].save(update_fields=["phone_number"])

        response = self.client.post(
            "/api/auth/register/",
            {
                "full_name": "Conta do Jogador",
                "phone_number": "(27) 98888-7777",
                "username": "conta_jogador",
                "password": "SenhaNova123",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201)
        created_user = User.objects.get(username="conta_jogador")
        self.assertIsNone(created_user.linked_player)
        self.assertEqual(Player.objects.filter(phone_number="27988887777").count(), 1)

    def test_current_user_can_update_profile(self) -> None:
        self.client.credentials(HTTP_AUTHORIZATION=f"Token {self.common_token.key}")

        response = self.client.patch(
            "/api/auth/me/",
            {
                "username": "comum_editado",
                "display_name": "Jogador Atualizado",
                "email": "jogador.atualizado@pelada.local",
                "phone_number": "(27) 97777-1111",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["username"], "comum_editado")
        self.assertEqual(response.data["display_name"], "Jogador Atualizado")
        self.assertEqual(response.data["phone_number"], "27977771111")

    def test_current_user_profile_rejects_username_with_spaces(self) -> None:
        self.client.credentials(HTTP_AUTHORIZATION=f"Token {self.common_token.key}")

        response = self.client.patch(
            "/api/auth/me/",
            {"username": "comum editado"},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("username", response.data)

    def test_admin_account_creation_rejects_username_with_spaces(self) -> None:
        response = self.client.post(
            "/api/users/",
            {
                "username": "novo admin",
                "email": "novo.admin@pelada.local",
                "display_name": "Novo Admin",
                "role": Role.ADMIN,
                "is_active": True,
                "must_change_password": True,
                "password": "SenhaNova123",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("username", response.data)

    def test_financial_summary_endpoint(self) -> None:
        response = self.client.get("/api/dashboard/financial-summary/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(Decimal(response.data["inflow_total"]), Decimal("300.00"))
        self.assertEqual(Decimal(response.data["outflow_total"]), Decimal("120.00"))
        self.assertEqual(Decimal(response.data["pending_total"]), Decimal("45.00"))
        self.assertEqual(Decimal(response.data["current_balance"]), Decimal("180.00"))

    def test_guest_creation_requires_invited_by_player(self) -> None:
        response = self.client.post(
            "/api/attendance/",
            {
                "match": self.match.id,
                "player": None,
                "display_name": "Convidado Sem Responsavel",
                "is_guest": True,
                "attendance_status": MatchAttendance.AttendanceStatus.CONFIRMED,
                "overall": 66,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("invited_by", response.data)

        created_response = self.client.post(
            "/api/attendance/",
            {
                "match": self.match.id,
                "player": None,
                "display_name": "Convidado Com Responsavel",
                "is_guest": True,
                "attendance_status": MatchAttendance.AttendanceStatus.CONFIRMED,
                "invited_by": self.players[0].id,
                "overall": 66,
            },
            format="json",
        )

        self.assertEqual(created_response.status_code, 201)
        self.assertEqual(str(created_response.data["invited_by"]), str(self.players[0].id))
        self.assertEqual(created_response.data["invited_by_name"], self.players[0].full_name)

    def test_guest_fee_is_registered_as_pending_after_match_ends_and_can_be_paid(self) -> None:
        self.match.status = Match.Status.ARCHIVED
        self.match.archived_at = timezone.now()
        self.match.save(update_fields=["status", "archived_at"])
        guest = MatchAttendance.objects.create(
            match=self.match,
            display_name="Convidado Devendo",
            is_guest=True,
            attendance_status=MatchAttendance.AttendanceStatus.CONFIRMED,
            overall=66,
            guest_fee_amount=Decimal("14.00"),
            guest_fee_status=MatchAttendance.GuestFeeStatus.PENDING,
        )

        summary_response = self.client.get("/api/dashboard/financial-summary/")
        attendance_response = self.client.get(f"/api/attendance/?match={self.match.id}")
        guest_fee_due_response = self.client.get("/api/attendance/?guest_fee_due=true")
        paid_response = self.client.post(f"/api/attendance/{guest.id}/mark-guest-fee-paid/")
        next_summary_response = self.client.get("/api/dashboard/financial-summary/")
        next_guest_fee_due_response = self.client.get("/api/attendance/?guest_fee_due=true")

        self.assertEqual(summary_response.status_code, 200)
        self.assertEqual(Decimal(summary_response.data["pending_total"]), Decimal("59.00"))
        self.assertEqual(guest_fee_due_response.status_code, 200)
        self.assertEqual([item["id"] for item in guest_fee_due_response.data], [str(guest.id)])
        guest_payload = next(item for item in attendance_response.data if item["id"] == str(guest.id))
        self.assertTrue(guest_payload["guest_fee_is_due"])
        self.assertEqual(Decimal(guest_payload["guest_fee_outstanding"]), Decimal("14.00"))
        self.assertEqual(paid_response.status_code, 200)
        self.assertEqual(paid_response.data["guest_fee_status"], MatchAttendance.GuestFeeStatus.PAID)
        self.assertFalse(paid_response.data["guest_fee_is_due"])
        self.assertEqual(Decimal(next_summary_response.data["pending_total"]), Decimal("45.00"))
        self.assertEqual(next_guest_fee_due_response.data, [])
        self.assertTrue(
            Transaction.objects.filter(
                external_reference=f"guest-fee:{guest.id}",
                amount=Decimal("14.00"),
                status=Transaction.Status.POSTED,
            ).exists()
        )

    def test_guest_fee_can_be_waived_without_posting_transaction(self) -> None:
        self.match.status = Match.Status.ARCHIVED
        self.match.archived_at = timezone.now()
        self.match.save(update_fields=["status", "archived_at"])
        guest = MatchAttendance.objects.create(
            match=self.match,
            display_name="Convidado Desconsiderado",
            is_guest=True,
            attendance_status=MatchAttendance.AttendanceStatus.CONFIRMED,
            invited_by=self.players[0],
            overall=66,
            guest_fee_amount=Decimal("14.00"),
            guest_fee_status=MatchAttendance.GuestFeeStatus.PENDING,
        )

        summary_response = self.client.get("/api/dashboard/financial-summary/")
        waive_response = self.client.post(f"/api/attendance/{guest.id}/waive-guest-fee/")
        next_summary_response = self.client.get("/api/dashboard/financial-summary/")
        next_guest_fee_due_response = self.client.get("/api/attendance/?guest_fee_due=true")

        self.assertEqual(summary_response.status_code, 200)
        self.assertEqual(Decimal(summary_response.data["pending_total"]), Decimal("59.00"))
        self.assertEqual(waive_response.status_code, 200)
        self.assertEqual(waive_response.data["guest_fee_status"], MatchAttendance.GuestFeeStatus.WAIVED)
        self.assertFalse(waive_response.data["guest_fee_is_due"])
        self.assertEqual(Decimal(next_summary_response.data["pending_total"]), Decimal("45.00"))
        self.assertEqual(next_guest_fee_due_response.data, [])
        self.assertFalse(
            Transaction.objects.filter(
                external_reference=f"guest-fee:{guest.id}",
            ).exists()
        )

    def test_generate_teams_action_assigns_attendance(self) -> None:
        stale_attendance = MatchAttendance.objects.create(
            match=self.match,
            display_name="Convidado antigo",
            is_guest=True,
            attendance_status=MatchAttendance.AttendanceStatus.DECLINED,
            assigned_team_number=3,
            assigned_team_name="Time antigo",
            overall=60,
        )

        response = self.client.post(
            f"/api/matches/{self.match.id}/generate-teams/",
            {"team_count": 2},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data["teams"]), 2)
        assigned_ids = [
            player["id"]
            for team in response.data["teams"]
            for player in team["players"]
        ]
        self.assertEqual(len(assigned_ids), len(set(assigned_ids)))
        self.match.refresh_from_db()
        self.assertIsNotNone(self.match.teams_generated_at)
        self.assertTrue(
            MatchAttendance.objects.filter(match=self.match, assigned_team_number__isnull=False).exists()
        )
        stale_attendance.refresh_from_db()
        self.assertIsNone(stale_attendance.assigned_team_number)
        self.assertEqual(stale_attendance.assigned_team_name, "")

    def test_clear_teams_action_removes_generated_assignments(self) -> None:
        self.client.post(
            f"/api/matches/{self.match.id}/generate-teams/",
            {"team_count": 2},
            format="json",
        )
        self.match.refresh_from_db()
        self.assertIsNotNone(self.match.teams_generated_at)

        response = self.client.post(f"/api/matches/{self.match.id}/clear-teams/")

        self.assertEqual(response.status_code, 200)
        self.assertIsNone(response.data["teams_generated_at"])
        self.match.refresh_from_db()
        self.assertIsNone(self.match.teams_generated_at)
        self.assertFalse(
            MatchAttendance.objects.filter(match=self.match, assigned_team_number__isnull=False).exists()
        )
        self.assertFalse(
            MatchAttendance.objects.filter(match=self.match).exclude(assigned_team_name="").exists()
        )

    def test_admin_can_swap_generated_team_players(self) -> None:
        self.client.post(
            f"/api/matches/{self.match.id}/generate-teams/",
            {"team_count": 2},
            format="json",
        )
        attendances = list(
            MatchAttendance.objects.filter(match=self.match)
            .order_by("assigned_team_number", "display_name")
        )
        source_attendance = attendances[0]
        target_attendance = next(
            attendance
            for attendance in attendances
            if attendance.assigned_team_number != source_attendance.assigned_team_number
        )
        source_team_number = source_attendance.assigned_team_number
        source_team_name = source_attendance.assigned_team_name
        target_team_number = target_attendance.assigned_team_number
        target_team_name = target_attendance.assigned_team_name

        response = self.client.post(
            f"/api/matches/{self.match.id}/swap-team-players/",
            {
                "source_attendance_id": str(source_attendance.id),
                "target_attendance_id": str(target_attendance.id),
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        source_attendance.refresh_from_db()
        target_attendance.refresh_from_db()
        self.assertEqual(source_attendance.assigned_team_number, target_team_number)
        self.assertEqual(source_attendance.assigned_team_name, target_team_name)
        self.assertEqual(target_attendance.assigned_team_number, source_team_number)
        self.assertEqual(target_attendance.assigned_team_name, source_team_name)
        response_assignments = {
            item["id"]: item["assigned_team_number"]
            for item in response.data
        }
        self.assertEqual(response_assignments[str(source_attendance.id)], target_team_number)
        self.assertEqual(response_assignments[str(target_attendance.id)], source_team_number)

    def test_admin_can_export_import_match_stats_and_update_sports_ranking(self) -> None:
        self.client.post(
            f"/api/matches/{self.match.id}/generate-teams/",
            {"team_count": 2},
            format="json",
        )
        attendances = list(
            MatchAttendance.objects.filter(match=self.match)
            .select_related("player")
            .order_by("assigned_team_number", "display_name")
        )
        target_attendance = attendances[0]
        assistant_attendance = attendances[1]
        winning_team_number = target_attendance.assigned_team_number
        winning_team_name = target_attendance.assigned_team_name

        export_response = self.client.get(f"/api/matches/{self.match.id}/stats-sheet/")
        self.assertEqual(export_response.status_code, 200)
        self.assertIn("spreadsheetml.sheet", export_response["Content-Type"])

        workbook = load_workbook(io.BytesIO(export_response.content))
        self.assertIn("Estatisticas", workbook.sheetnames)
        worksheet = workbook["Estatisticas"]
        visible_values = [
            str(cell.value)
            for row in worksheet.iter_rows()
            for cell in row
            if cell.value is not None
        ]
        self.assertNotIn(str(target_attendance.id), visible_values)
        self.assertEqual(worksheet["A4"].fill.fgColor.rgb, "00EDE9FE")
        self.assertGreater(worksheet.column_dimensions["A"].width, 25)

        winning_team_row = None
        rows_by_player_name = {}
        for row_number in range(1, worksheet.max_row + 1):
            first_value = worksheet.cell(row_number, 1).value
            third_value = worksheet.cell(row_number, 3).value
            if first_value == winning_team_name and third_value == "Vitoria do time":
                winning_team_row = row_number
            if first_value in {attendance.display_name for attendance in attendances}:
                rows_by_player_name[first_value] = row_number

        self.assertIsNotNone(winning_team_row)
        worksheet.cell(winning_team_row, 4, "SIM")
        for attendance in attendances:
            row_number = rows_by_player_name[attendance.display_name]
            worksheet.cell(row_number, 3, 2 if attendance.id == target_attendance.id else 0)
            worksheet.cell(
                row_number,
                4,
                1 if attendance.id == target_attendance.id else 2 if attendance.id == assistant_attendance.id else 0,
            )

        output = io.BytesIO()
        workbook.save(output)
        uploaded_file = SimpleUploadedFile(
            "estatisticas.xlsx",
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        import_response = self.client.post(
            f"/api/matches/{self.match.id}/import-stats-sheet/",
            {"file": uploaded_file},
            format="multipart",
        )
        ranking_response = self.client.get("/api/analytics/sports-ranking/?limit=10")

        self.assertEqual(import_response.status_code, 200)
        self.assertEqual(import_response.data["players_processed"], len(attendances))
        self.assertEqual(import_response.data["goals_total"], 2)
        self.assertEqual(import_response.data["assists_total"], 3)
        self.assertEqual(MatchPlayerStat.objects.filter(match=self.match).count(), len(attendances))
        self.assertTrue(
            MatchPlayerStat.objects.filter(
                match=self.match,
                team_number=winning_team_number,
                team_won=True,
            ).exists()
        )
        self.assertEqual(ranking_response.status_code, 200)
        self.assertEqual(ranking_response.data["top_scorers"][0]["player_id"], str(target_attendance.player_id))
        self.assertEqual(ranking_response.data["top_scorers"][0]["goals"], 2)
        self.assertEqual(ranking_response.data["top_assistants"][0]["player_id"], str(assistant_attendance.player_id))
        self.assertEqual(ranking_response.data["top_assistants"][0]["assists"], 2)

        self.client.credentials(HTTP_AUTHORIZATION=f"Token {self.common_token.key}")
        common_ranking_response = self.client.get("/api/analytics/sports-ranking/?limit=10")
        common_export_response = self.client.get(f"/api/matches/{self.match.id}/stats-sheet/")
        self.assertEqual(common_ranking_response.status_code, 200)
        self.assertEqual(common_export_response.status_code, 403)

    def test_overall_history_lists_member_snapshots_for_all_authenticated_users(self) -> None:
        self.match.status = Match.Status.ARCHIVED
        self.match.scheduled_at = timezone.make_aware(datetime(2026, 5, 26, 20, 0))
        self.match.save(update_fields=["status", "scheduled_at"])
        old_match = Match.objects.create(
            scheduled_at=timezone.make_aware(datetime(2026, 5, 25, 20, 0)),
            status=Match.Status.ARCHIVED,
            expected_team_count=2,
            created_by=self.user,
        )
        excluded_match = Match.objects.create(
            scheduled_at=timezone.make_aware(datetime(2026, 6, 30, 20, 0)),
            status=Match.Status.ARCHIVED,
            expected_team_count=2,
            created_by=self.user,
        )
        MatchAttendance.objects.create(
            match=old_match,
            player=self.players[0],
            display_name=self.players[0].full_name,
            is_guest=False,
            attendance_status=MatchAttendance.AttendanceStatus.CONFIRMED,
            overall=55,
        )
        MatchAttendance.objects.create(
            match=excluded_match,
            player=self.players[0],
            display_name=self.players[0].full_name,
            is_guest=False,
            attendance_status=MatchAttendance.AttendanceStatus.CONFIRMED,
            overall=77,
        )
        MatchAttendance.objects.create(
            match=self.match,
            display_name="Convidado Historico",
            is_guest=True,
            attendance_status=MatchAttendance.AttendanceStatus.CONFIRMED,
            overall=66,
        )
        self.players[0].overall = 88
        self.players[0].save(update_fields=["overall"])

        admin_response = self.client.get("/api/matches/overall-history/")
        self.client.credentials(HTTP_AUTHORIZATION=f"Token {self.common_token.key}")
        common_response = self.client.get("/api/matches/overall-history/")

        self.assertEqual(admin_response.status_code, 200)
        self.assertEqual(common_response.status_code, 200)
        self.assertIn(str(self.players[0].id), [item["player_id"] for item in admin_response.data["players"]])
        self.assertEqual(len(admin_response.data["matches"]), 1)
        history_points = admin_response.data["matches"][0]["points"]
        self.assertIn(str(self.players[0].id), [item["player_id"] for item in history_points])
        player_point = next(item for item in history_points if item["player_id"] == str(self.players[0].id))
        self.assertEqual(player_point["overall"], 88)
        self.assertNotIn("Convidado Historico", [item["display_name"] for item in history_points])

    def test_admin_can_create_match_via_api(self) -> None:
        response = self.client.post(
            "/api/matches/",
            {
                "scheduled_at": (timezone.now() + timedelta(days=9)).isoformat(),
                "location": "Arena Roxa",
                "status": Match.Status.OPEN,
                "expected_team_count": 3,
                "notes": "Rodada extra do feriado",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201)
        created_match = Match.objects.get(location="Arena Roxa")
        self.assertFalse(MatchAttendance.objects.filter(match=created_match).exists())

    def test_admin_can_patch_match_status(self) -> None:
        response = self.client.patch(
            f"/api/matches/{self.match.id}/",
            {"status": Match.Status.ARCHIVED},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.match.refresh_from_db()
        self.assertEqual(self.match.status, Match.Status.ARCHIVED)
        self.assertIsNotNone(self.match.attendance_locked_at)

    def test_admin_can_finalize_match_with_winning_team_and_update_wins_ranking(self) -> None:
        self.client.post(
            f"/api/matches/{self.match.id}/generate-teams/",
            {"team_count": 2},
            format="json",
        )
        attendances = list(MatchAttendance.objects.filter(match=self.match))
        winning_team_number = attendances[0].assigned_team_number
        winners = [
            attendance
            for attendance in attendances
            if attendance.assigned_team_number == winning_team_number
        ]

        response = self.client.post(
            f"/api/matches/{self.match.id}/finalize/",
            {"winning_team_number": winning_team_number},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.match.refresh_from_db()
        self.assertEqual(self.match.status, Match.Status.ARCHIVED)
        self.assertEqual(self.match.winning_team_number, winning_team_number)
        self.assertIsNotNone(self.match.attendance_locked_at)
        self.assertIsNotNone(self.match.archived_at)
        self.assertEqual(
            MatchPlayerStat.objects.filter(match=self.match, team_won=True).count(),
            len(winners),
        )
        self.assertFalse(
            MatchPlayerStat.objects.filter(match=self.match, team_won=True)
            .exclude(team_number=winning_team_number)
            .exists()
        )

        ranking_response = self.client.get("/api/analytics/sports-ranking/?limit=10")
        self.assertEqual(ranking_response.status_code, 200)
        winner_ids = {str(attendance.player_id) for attendance in winners}
        ranked_winner_ids = {entry["player_id"] for entry in ranking_response.data["top_winners"]}
        self.assertEqual(ranked_winner_ids, winner_ids)
        self.assertTrue(all(entry["wins"] == 1 for entry in ranking_response.data["top_winners"]))

    def test_finalize_match_without_winner_keeps_wins_untouched(self) -> None:
        response = self.client.post(
            f"/api/matches/{self.match.id}/finalize/",
            {},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.match.refresh_from_db()
        self.assertEqual(self.match.status, Match.Status.ARCHIVED)
        self.assertIsNone(self.match.winning_team_number)
        self.assertFalse(MatchPlayerStat.objects.filter(match=self.match).exists())

    def test_finalize_match_rejects_team_outside_the_match(self) -> None:
        self.client.post(
            f"/api/matches/{self.match.id}/generate-teams/",
            {"team_count": 2},
            format="json",
        )

        response = self.client.post(
            f"/api/matches/{self.match.id}/finalize/",
            {"winning_team_number": 9},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.match.refresh_from_db()
        self.assertEqual(self.match.status, Match.Status.OPEN)
        self.assertIsNone(self.match.winning_team_number)

    def test_common_user_cannot_finalize_match(self) -> None:
        self.client.credentials(HTTP_AUTHORIZATION=f"Token {self.common_token.key}")

        response = self.client.post(
            f"/api/matches/{self.match.id}/finalize/",
            {},
            format="json",
        )

        self.assertEqual(response.status_code, 403)
        self.match.refresh_from_db()
        self.assertEqual(self.match.status, Match.Status.OPEN)

    def test_admin_can_record_match_result(self) -> None:
        response = self.client.patch(
            f"/api/matches/{self.match.id}/",
            {"result_summary": "Time Roxo 7 x 5 Time Cinza"},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.match.refresh_from_db()
        self.assertEqual(self.match.result_summary, "Time Roxo 7 x 5 Time Cinza")
        self.assertIsNotNone(self.match.result_recorded_at)

    def test_admin_can_link_user_to_player(self) -> None:
        response = self.client.patch(
            f"/api/users/{self.common_user.id}/",
            {"linked_player": str(self.players[0].id)},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.common_user.refresh_from_db()
        self.assertEqual(self.common_user.linked_player_id, self.players[0].id)

    def test_admin_can_list_users(self) -> None:
        response = self.client.get("/api/users/")

        self.assertEqual(response.status_code, 200)
        self.assertGreaterEqual(len(response.data), 2)

    def test_admin_can_create_user_without_password(self) -> None:
        response = self.client.post(
            "/api/users/",
            {
                "username": "novo-comum",
                "email": "novo@pelada.local",
                "display_name": "Novo Comum",
                "role": Role.COMMON,
                "linked_player": str(self.players[1].id),
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201)
        created_user = User.objects.get(username="novo-comum")
        self.assertFalse(created_user.has_usable_password())
        self.assertTrue(created_user.must_change_password)
        self.assertEqual(created_user.linked_player_id, self.players[1].id)

    def test_admin_can_reset_user_password(self) -> None:
        self.common_user.must_change_password = False
        self.common_user.save(update_fields=["must_change_password"])

        response = self.client.post(
            f"/api/users/{self.common_user.id}/reset-password/",
            {"new_password": "NovaSenhaSegura@123"},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.common_user.refresh_from_db()
        self.assertTrue(self.common_user.check_password("NovaSenhaSegura@123"))
        self.assertTrue(self.common_user.must_change_password)

    def test_common_user_can_change_own_password(self) -> None:
        self.common_user.must_change_password = True
        self.common_user.set_password("SenhaTemporaria@123")
        self.common_user.save(update_fields=["must_change_password", "password"])
        self.client.credentials(HTTP_AUTHORIZATION=f"Token {self.common_token.key}")

        response = self.client.post(
            "/api/auth/change-password/",
            {
                "current_password": "SenhaTemporaria@123",
                "new_password": "SenhaDefinitiva@123",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn("token", response.data)
        self.common_user.refresh_from_db()
        self.assertTrue(self.common_user.check_password("SenhaDefinitiva@123"))
        self.assertFalse(self.common_user.must_change_password)

    def test_change_password_rejects_invalid_current_password(self) -> None:
        self.client.credentials(HTTP_AUTHORIZATION=f"Token {self.common_token.key}")

        response = self.client.post(
            "/api/auth/change-password/",
            {
                "current_password": "senha-incorreta",
                "new_password": "SenhaNova@999",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("current_password", response.data)

    def test_common_user_cannot_access_user_management(self) -> None:
        self.client.credentials(HTTP_AUTHORIZATION=f"Token {self.common_token.key}")

        response = self.client.get("/api/users/")

        self.assertEqual(response.status_code, 403)

    def test_common_user_portal_overview(self) -> None:
        self.common_user.linked_player = self.players[0]
        self.common_user.save(update_fields=["linked_player"])
        Transaction.objects.create(
            direction=Transaction.Direction.INFLOW,
            category=Transaction.Category.MONTHLY_FEE,
            status=Transaction.Status.POSTED,
            amount=Decimal("70.00"),
            description="Mensalidade jogador comum",
            occurred_on=date.today(),
            related_player=self.players[0],
            recorded_by=self.user,
        )
        self.client.credentials(HTTP_AUTHORIZATION=f"Token {self.common_token.key}")

        response = self.client.get("/api/portal/me/overview/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["user"]["id"], str(self.common_user.id))
        self.assertEqual(response.data["linked_player"]["id"], str(self.players[0].id))
        self.assertEqual(Decimal(response.data["financial_status"]["paid_amount"]), Decimal("70.00"))
        self.assertGreaterEqual(response.data["attendance_status"]["confirmed_count"], 1)

    def test_admin_can_access_analytics_endpoints(self) -> None:
        season_response = self.client.get("/api/analytics/season-overview/")
        presence_response = self.client.get("/api/analytics/presence-ranking/?limit=3")
        payment_response = self.client.get("/api/analytics/payment-ranking/?limit=3")

        self.assertEqual(season_response.status_code, 200)
        self.assertEqual(presence_response.status_code, 200)
        self.assertEqual(payment_response.status_code, 200)
        self.assertEqual(season_response.data["active_members"], 4)
        self.assertGreaterEqual(len(presence_response.data["ranking"]), 1)
        self.assertGreaterEqual(len(payment_response.data["ranking"]), 1)

    def test_common_user_cannot_access_admin_analytics(self) -> None:
        self.client.credentials(HTTP_AUTHORIZATION=f"Token {self.common_token.key}")

        response = self.client.get("/api/analytics/season-overview/")

        self.assertEqual(response.status_code, 403)

    def test_common_user_reads_only_own_extract(self) -> None:
        self.common_user.linked_player = self.players[0]
        self.common_user.save(update_fields=["linked_player"])
        own_transaction = Transaction.objects.create(
            direction=Transaction.Direction.INFLOW,
            category=Transaction.Category.MONTHLY_FEE,
            status=Transaction.Status.POSTED,
            amount=Decimal("120.00"),
            description="Mensalidade propria",
            occurred_on=date.today(),
            related_player=self.players[0],
            recorded_by=self.user,
        )
        Transaction.objects.create(
            direction=Transaction.Direction.INFLOW,
            category=Transaction.Category.MONTHLY_FEE,
            status=Transaction.Status.POSTED,
            amount=Decimal("120.00"),
            description="Mensalidade outro jogador",
            occurred_on=date.today(),
            related_player=self.players[1],
            recorded_by=self.user,
        )
        self.client.credentials(HTTP_AUTHORIZATION=f"Token {self.common_token.key}")

        response = self.client.get("/api/transactions/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual([item["id"] for item in response.data], [str(own_transaction.id)])

    def test_common_user_reads_active_roster_and_matches(self) -> None:
        self.players[1].is_active = False
        self.players[1].save(update_fields=["is_active"])
        archived_match = Match.objects.create(
            scheduled_at=timezone.now() + timedelta(days=4),
            status=Match.Status.ARCHIVED,
            expected_team_count=2,
            created_by=self.user,
        )
        self.client.credentials(HTTP_AUTHORIZATION=f"Token {self.common_token.key}")

        players_response = self.client.get("/api/players/")
        matches_response = self.client.get("/api/matches/")
        summary_response = self.client.get("/api/dashboard/financial-summary/")

        self.assertEqual(players_response.status_code, 200)
        self.assertEqual(matches_response.status_code, 200)
        self.assertEqual(summary_response.status_code, 403)
        self.assertNotIn(str(self.players[1].id), [item["id"] for item in players_response.data])
        self.assertIn(str(archived_match.id), [item["id"] for item in matches_response.data])
        self.assertTrue(
            all(
                item["status"] in {Match.Status.OPEN, Match.Status.ARCHIVED}
                for item in matches_response.data
            )
        )

    def test_common_user_without_linked_player_cannot_rate_archived_match(self) -> None:
        self.match.status = Match.Status.ARCHIVED
        self.match.archived_at = timezone.now()
        self.match.save(update_fields=["status", "archived_at"])
        target_attendance = MatchAttendance.objects.get(match=self.match, player=self.players[1])
        original_overall = self.players[1].overall
        self.client.credentials(HTTP_AUTHORIZATION=f"Token {self.common_token.key}")

        state_response = self.client.get(f"/api/matches/{self.match.id}/player-ratings/")
        submit_response = self.client.post(
            f"/api/matches/{self.match.id}/player-ratings/",
            {"ratings": [{"attendance_id": str(target_attendance.id), "score": 6.5}]},
            format="json",
        )

        self.assertEqual(state_response.status_code, 200)
        self.assertFalse(state_response.data["can_rate"])
        self.assertIn("Procure um administrador", state_response.data["locked_reason"])
        self.assertEqual(submit_response.status_code, 400)
        self.assertEqual(MatchPlayerRating.objects.filter(match=self.match, rater_user=self.common_user).count(), 0)
        self.players[1].refresh_from_db()
        self.assertEqual(self.players[1].overall, original_overall)

    def test_common_user_can_rate_only_teammates_and_not_self_or_opponents(self) -> None:
        self.common_user.linked_player = self.players[0]
        self.common_user.save(update_fields=["linked_player"])
        other_match = Match.objects.create(
            scheduled_at=timezone.now() - timedelta(days=1),
            status=Match.Status.ARCHIVED,
            archived_at=timezone.now(),
            expected_team_count=2,
            created_by=self.user,
        )
        self_attendance = MatchAttendance.objects.create(
            match=other_match,
            player=self.players[0],
            display_name=self.players[0].full_name,
            is_guest=False,
            attendance_status=MatchAttendance.AttendanceStatus.CONFIRMED,
            assigned_team_number=1,
            assigned_team_name="Time 1",
            overall=self.players[0].overall,
        )
        teammate_attendance = MatchAttendance.objects.create(
            match=other_match,
            player=self.players[1],
            display_name=self.players[1].full_name,
            is_guest=False,
            attendance_status=MatchAttendance.AttendanceStatus.CONFIRMED,
            assigned_team_number=1,
            assigned_team_name="Time 1",
            overall=self.players[1].overall,
        )
        opponent_attendance = MatchAttendance.objects.create(
            match=other_match,
            player=self.players[2],
            display_name=self.players[2].full_name,
            is_guest=False,
            attendance_status=MatchAttendance.AttendanceStatus.CONFIRMED,
            assigned_team_number=2,
            assigned_team_name="Time 2",
            overall=self.players[2].overall,
        )
        self.client.credentials(HTTP_AUTHORIZATION=f"Token {self.common_token.key}")

        state_response = self.client.get(f"/api/matches/{other_match.id}/player-ratings/")
        response = self.client.post(
            f"/api/matches/{other_match.id}/player-ratings/",
            {"ratings": [{"attendance_id": str(teammate_attendance.id), "score": 10}]},
            format="json",
        )
        opponent_response = self.client.post(
            f"/api/matches/{other_match.id}/player-ratings/",
            {"ratings": [{"attendance_id": str(opponent_attendance.id), "score": 10}]},
            format="json",
        )
        self_response = self.client.post(
            f"/api/matches/{other_match.id}/player-ratings/",
            {"ratings": [{"attendance_id": str(self_attendance.id), "score": 10}]},
            format="json",
        )

        self.assertEqual(state_response.status_code, 200)
        self.assertTrue(state_response.data["can_rate"])
        self.assertEqual(
            [item["attendance_id"] for item in state_response.data["items"]],
            [str(teammate_attendance.id)],
        )
        self.assertNotIn(str(self_attendance.id), [item["attendance_id"] for item in state_response.data["items"]])
        self.assertEqual(response.status_code, 200)
        self.assertEqual(MatchPlayerRating.objects.filter(match=other_match, rater_user=self.common_user).count(), 1)
        self.assertEqual(opponent_response.status_code, 400)
        self.assertEqual(self_response.status_code, 400)

    def test_rater_can_skip_a_player_and_skipping_removes_a_previous_vote(self) -> None:
        self.common_user.linked_player = self.players[0]
        self.common_user.save(update_fields=["linked_player"])
        self.match.status = Match.Status.ARCHIVED
        self.match.archived_at = timezone.now()
        self.match.rating_mode = Match.RatingMode.GENERAL
        self.match.save(update_fields=["status", "archived_at", "rating_mode"])
        MatchAttendance.objects.filter(match=self.match).update(assigned_team_number=None)
        rated_attendance = MatchAttendance.objects.get(match=self.match, player=self.players[1])
        skipped_attendance = MatchAttendance.objects.get(match=self.match, player=self.players[2])
        self.client.credentials(HTTP_AUTHORIZATION=f"Token {self.common_token.key}")

        skip_response = self.client.post(
            f"/api/matches/{self.match.id}/player-ratings/",
            {
                "ratings": [
                    {"attendance_id": str(rated_attendance.id), "score": 8},
                    {"attendance_id": str(skipped_attendance.id), "score": None},
                ]
            },
            format="json",
        )
        rated_ids_after_skip = set(
            MatchPlayerRating.objects.filter(
                match=self.match, rater_user=self.common_user
            ).values_list("rated_attendance_id", flat=True)
        )

        vote_response = self.client.post(
            f"/api/matches/{self.match.id}/player-ratings/",
            {"ratings": [{"attendance_id": str(skipped_attendance.id), "score": 9}]},
            format="json",
        )
        unvote_response = self.client.post(
            f"/api/matches/{self.match.id}/player-ratings/",
            {
                "ratings": [
                    {"attendance_id": str(rated_attendance.id), "score": 8},
                    {"attendance_id": str(skipped_attendance.id), "score": None},
                ]
            },
            format="json",
        )
        rated_ids_after_unvote = set(
            MatchPlayerRating.objects.filter(
                match=self.match, rater_user=self.common_user
            ).values_list("rated_attendance_id", flat=True)
        )

        self.assertEqual(skip_response.status_code, 200)
        self.assertEqual(rated_ids_after_skip, {rated_attendance.id})
        self.assertEqual(vote_response.status_code, 200)
        self.assertEqual(unvote_response.status_code, 200)
        self.assertEqual(rated_ids_after_unvote, {rated_attendance.id})

    def test_general_rating_mode_lets_player_rate_everyone_except_themselves(self) -> None:
        self.common_user.linked_player = self.players[0]
        self.common_user.save(update_fields=["linked_player"])
        general_match = Match.objects.create(
            scheduled_at=timezone.now() - timedelta(days=1),
            status=Match.Status.ARCHIVED,
            archived_at=timezone.now(),
            expected_team_count=2,
            rating_mode=Match.RatingMode.GENERAL,
            created_by=self.user,
        )
        self_attendance = MatchAttendance.objects.create(
            match=general_match,
            player=self.players[0],
            display_name=self.players[0].full_name,
            is_guest=False,
            attendance_status=MatchAttendance.AttendanceStatus.CONFIRMED,
            assigned_team_number=1,
            assigned_team_name="Time 1",
            overall=self.players[0].overall,
        )
        teammate_attendance = MatchAttendance.objects.create(
            match=general_match,
            player=self.players[1],
            display_name=self.players[1].full_name,
            is_guest=False,
            attendance_status=MatchAttendance.AttendanceStatus.CONFIRMED,
            assigned_team_number=1,
            assigned_team_name="Time 1",
            overall=self.players[1].overall,
        )
        opponent_attendance = MatchAttendance.objects.create(
            match=general_match,
            player=self.players[2],
            display_name=self.players[2].full_name,
            is_guest=False,
            attendance_status=MatchAttendance.AttendanceStatus.CONFIRMED,
            assigned_team_number=2,
            assigned_team_name="Time 2",
            overall=self.players[2].overall,
        )
        self.client.credentials(HTTP_AUTHORIZATION=f"Token {self.common_token.key}")

        state_response = self.client.get(f"/api/matches/{general_match.id}/player-ratings/")
        opponent_response = self.client.post(
            f"/api/matches/{general_match.id}/player-ratings/",
            {"ratings": [{"attendance_id": str(opponent_attendance.id), "score": 9}]},
            format="json",
        )
        self_response = self.client.post(
            f"/api/matches/{general_match.id}/player-ratings/",
            {"ratings": [{"attendance_id": str(self_attendance.id), "score": 10}]},
            format="json",
        )

        self.assertEqual(state_response.status_code, 200)
        self.assertTrue(state_response.data["can_rate"])
        rateable_ids = {item["attendance_id"] for item in state_response.data["items"]}
        self.assertEqual(
            rateable_ids,
            {str(teammate_attendance.id), str(opponent_attendance.id)},
        )
        self.assertNotIn(str(self_attendance.id), rateable_ids)
        self.assertEqual(opponent_response.status_code, 200)
        self.assertEqual(self_response.status_code, 400)

    def test_general_rating_mode_does_not_require_a_generated_team(self) -> None:
        self.common_user.linked_player = self.players[0]
        self.common_user.save(update_fields=["linked_player"])
        self.match.status = Match.Status.ARCHIVED
        self.match.archived_at = timezone.now()
        self.match.rating_mode = Match.RatingMode.GENERAL
        self.match.save(update_fields=["status", "archived_at", "rating_mode"])
        MatchAttendance.objects.filter(match=self.match).update(assigned_team_number=None)
        target_attendance = MatchAttendance.objects.get(match=self.match, player=self.players[1])
        self.client.credentials(HTTP_AUTHORIZATION=f"Token {self.common_token.key}")

        state_response = self.client.get(f"/api/matches/{self.match.id}/player-ratings/")
        submit_response = self.client.post(
            f"/api/matches/{self.match.id}/player-ratings/",
            {"ratings": [{"attendance_id": str(target_attendance.id), "score": 8}]},
            format="json",
        )

        self.assertEqual(state_response.status_code, 200)
        self.assertTrue(state_response.data["can_rate"])
        self.assertEqual(state_response.data["locked_reason"], "")
        self.assertEqual(submit_response.status_code, 200)
        self.assertEqual(
            MatchPlayerRating.objects.filter(match=self.match, rater_user=self.common_user).count(),
            1,
        )

    def test_matches_default_to_team_rating_mode(self) -> None:
        response = self.client.post(
            "/api/matches/",
            {
                "scheduled_at": (timezone.now() + timedelta(days=3)).isoformat(),
                "location": "Arena Padrao",
                "expected_team_count": 2,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["rating_mode"], Match.RatingMode.TEAM)

    def test_common_user_with_linked_player_without_generated_team_cannot_rate(self) -> None:
        self.common_user.linked_player = self.players[0]
        self.common_user.save(update_fields=["linked_player"])
        self.match.status = Match.Status.ARCHIVED
        self.match.archived_at = timezone.now()
        self.match.save(update_fields=["status", "archived_at"])
        target_attendance = MatchAttendance.objects.get(match=self.match, player=self.players[1])
        self.client.credentials(HTTP_AUTHORIZATION=f"Token {self.common_token.key}")

        state_response = self.client.get(f"/api/matches/{self.match.id}/player-ratings/")
        submit_response = self.client.post(
            f"/api/matches/{self.match.id}/player-ratings/",
            {"ratings": [{"attendance_id": str(target_attendance.id), "score": 8}]},
            format="json",
        )

        self.assertEqual(state_response.status_code, 200)
        self.assertFalse(state_response.data["can_rate"])
        self.assertIn("time gerado", state_response.data["locked_reason"])
        self.assertEqual(submit_response.status_code, 400)

    def test_rating_window_finalizes_overall_and_returns_empty_state_after_24_hours(self) -> None:
        self.match.status = Match.Status.ARCHIVED
        self.match.archived_at = timezone.now() - timedelta(hours=25)
        self.match.save(update_fields=["status", "archived_at"])
        target_attendance = MatchAttendance.objects.get(match=self.match, player=self.players[1])
        original_overall = self.players[1].overall
        MatchPlayerRating.objects.create(
            match=self.match,
            rater_user=self.common_user,
            rated_attendance=target_attendance,
            rated_player=self.players[1],
            score=10,
        )
        self.client.credentials(HTTP_AUTHORIZATION=f"Token {self.common_token.key}")

        response = self.client.get(f"/api/matches/{self.match.id}/player-ratings/")

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.data["can_rate"])
        self.assertEqual(response.data["items"], [])
        self.assertEqual(len(response.data["log"]), 1)
        self.assertEqual(response.data["log"][0]["rated_display_name"], target_attendance.display_name)
        self.assertEqual(response.data["log"][0]["score"], "10.0")
        self.assertIsNotNone(response.data["ratings_finalized_at"])
        summary_entry = next(
            item for item in response.data["overall_summary"] if item["player_id"] == str(self.players[1].id)
        )
        self.assertEqual(summary_entry["previous_overall"], original_overall)
        self.players[1].refresh_from_db()
        self.assertGreater(self.players[1].overall, original_overall)
        self.assertEqual(summary_entry["current_overall"], self.players[1].overall)
        self.assertGreater(summary_entry["delta"], 0)

    def test_admin_can_finalize_rating_window_manually(self) -> None:
        self.match.status = Match.Status.ARCHIVED
        self.match.archived_at = timezone.now()
        self.match.save(update_fields=["status", "archived_at"])
        target_attendance = MatchAttendance.objects.get(match=self.match, player=self.players[1])
        original_overall = self.players[1].overall
        MatchPlayerRating.objects.create(
            match=self.match,
            rater_user=self.common_user,
            rated_attendance=target_attendance,
            rated_player=self.players[1],
            score=10,
        )

        response = self.client.post(f"/api/matches/{self.match.id}/finalize-ratings/")

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.data["can_rate"])
        self.assertIsNotNone(response.data["ratings_finalized_at"])
        summary_entry = next(
            item for item in response.data["overall_summary"] if item["player_id"] == str(self.players[1].id)
        )
        self.players[1].refresh_from_db()
        self.assertEqual(summary_entry["previous_overall"], original_overall)
        self.assertEqual(summary_entry["current_overall"], self.players[1].overall)
        self.assertGreater(summary_entry["delta"], 0)

        self.client.credentials(HTTP_AUTHORIZATION=f"Token {self.common_token.key}")
        locked_response = self.client.post(
            f"/api/matches/{self.match.id}/player-ratings/",
            {"ratings": [{"attendance_id": str(target_attendance.id), "score": 1}]},
            format="json",
        )
        self.assertEqual(locked_response.status_code, 400)

    def test_rating_finalization_uses_only_match_average_with_25_percent_weight(self) -> None:
        target_player = self.players[1]
        target_player.overall = 70
        target_player.save(update_fields=["overall"])
        target_attendance = MatchAttendance.objects.get(match=self.match, player=target_player)
        target_attendance.overall = 70
        target_attendance.save(update_fields=["overall"])

        old_match = Match.objects.create(
            scheduled_at=timezone.now() - timedelta(days=7),
            status=Match.Status.ARCHIVED,
            archived_at=timezone.now() - timedelta(days=6),
            expected_team_count=2,
            created_by=self.user,
            ratings_finalized_at=timezone.now() - timedelta(days=5),
        )
        old_attendance = MatchAttendance.objects.create(
            match=old_match,
            player=target_player,
            display_name=target_player.full_name,
            is_guest=False,
            attendance_status=MatchAttendance.AttendanceStatus.CONFIRMED,
            overall=70,
        )
        second_rater = User.objects.create_user(
            username="comum-2",
            email="comum-2@pelada.local",
            password=self.password,
            role=Role.COMMON,
        )
        third_rater = User.objects.create_user(
            username="comum-3",
            email="comum-3@pelada.local",
            password=self.password,
            role=Role.COMMON,
        )
        MatchPlayerRating.objects.create(
            match=old_match,
            rater_user=self.common_user,
            rated_attendance=old_attendance,
            rated_player=target_player,
            score=1,
        )
        MatchPlayerRating.objects.create(
            match=old_match,
            rater_user=second_rater,
            rated_attendance=old_attendance,
            rated_player=target_player,
            score=1,
        )

        self.match.status = Match.Status.ARCHIVED
        self.match.archived_at = timezone.now()
        self.match.save(update_fields=["status", "archived_at"])
        for rater, score in (
            (self.common_user, Decimal("7.0")),
            (second_rater, Decimal("8.0")),
            (third_rater, Decimal("10.0")),
        ):
            MatchPlayerRating.objects.create(
                match=self.match,
                rater_user=rater,
                rated_attendance=target_attendance,
                rated_player=target_player,
                score=score,
            )

        response = self.client.post(f"/api/matches/{self.match.id}/finalize-ratings/")

        self.assertEqual(response.status_code, 200)
        target_player.refresh_from_db()
        self.assertEqual(target_player.overall, 77)
        summary_entry = next(
            item for item in response.data["overall_summary"] if item["player_id"] == str(target_player.id)
        )
        self.assertEqual(summary_entry["previous_overall"], 70)
        self.assertEqual(summary_entry["current_overall"], 77)
        self.assertEqual(summary_entry["delta"], 7)
        self.assertEqual(summary_entry["average_score"], "8.33")

    def test_rating_finalization_uses_all_votes_and_rounds_final_score_up(self) -> None:
        target_player = self.players[1]
        target_player.overall = 60
        target_player.save(update_fields=["overall"])
        target_attendance = MatchAttendance.objects.get(match=self.match, player=target_player)
        target_attendance.overall = 60
        target_attendance.save(update_fields=["overall"])
        raters = [
            self.common_user,
            User.objects.create_user(
                username="comum-2",
                email="comum-2@pelada.local",
                password=self.password,
                role=Role.COMMON,
            ),
            User.objects.create_user(
                username="comum-3",
                email="comum-3@pelada.local",
                password=self.password,
                role=Role.COMMON,
            ),
            User.objects.create_user(
                username="comum-4",
                email="comum-4@pelada.local",
                password=self.password,
                role=Role.COMMON,
            ),
        ]
        self.match.status = Match.Status.ARCHIVED
        self.match.archived_at = timezone.now()
        self.match.save(update_fields=["status", "archived_at"])
        scores = (Decimal("1.0"), Decimal("6.1"), Decimal("6.5"), Decimal("10.0"))
        for rater, score in zip(raters, scores):
            MatchPlayerRating.objects.create(
                match=self.match,
                rater_user=rater,
                rated_attendance=target_attendance,
                rated_player=target_player,
                score=score,
            )

        response = self.client.post(f"/api/matches/{self.match.id}/finalize-ratings/")

        self.assertEqual(response.status_code, 200)
        target_player.refresh_from_db()
        self.assertEqual(target_player.overall, 60)
        summary_entry = next(
            item for item in response.data["overall_summary"] if item["player_id"] == str(target_player.id)
        )
        self.assertEqual(summary_entry["average_score"], "5.90")
        self.assertEqual(summary_entry["rating_count"], 4)
        self.assertEqual(summary_entry["current_overall"], 60)

    def test_excellent_match_rating_can_move_high_overall_player_more_than_two_points(self) -> None:
        target_player = self.players[1]
        target_player.overall = 90
        target_player.save(update_fields=["overall"])
        target_attendance = MatchAttendance.objects.get(match=self.match, player=target_player)
        target_attendance.overall = 90
        target_attendance.save(update_fields=["overall"])
        second_rater = User.objects.create_user(
            username="comum-2",
            email="comum-2@pelada.local",
            password=self.password,
            role=Role.COMMON,
        )
        third_rater = User.objects.create_user(
            username="comum-3",
            email="comum-3@pelada.local",
            password=self.password,
            role=Role.COMMON,
        )
        self.match.status = Match.Status.ARCHIVED
        self.match.archived_at = timezone.now()
        self.match.save(update_fields=["status", "archived_at"])
        for rater in (self.common_user, second_rater, third_rater):
            MatchPlayerRating.objects.create(
                match=self.match,
                rater_user=rater,
                rated_attendance=target_attendance,
                rated_player=target_player,
                score=10,
            )

        response = self.client.post(f"/api/matches/{self.match.id}/finalize-ratings/")

        self.assertEqual(response.status_code, 200)
        target_player.refresh_from_db()
        self.assertEqual(target_player.overall, 95)
        summary_entry = next(
            item for item in response.data["overall_summary"] if item["player_id"] == str(target_player.id)
        )
        self.assertEqual(summary_entry["previous_overall"], 90)
        self.assertEqual(summary_entry["current_overall"], 95)
        self.assertEqual(summary_entry["delta"], 5)

    def test_rating_performance_bonus_starts_at_eight_and_increases_at_nine(self) -> None:
        first_player = self.players[1]
        first_player.overall = 80
        first_player.save(update_fields=["overall"])
        first_attendance = MatchAttendance.objects.get(match=self.match, player=first_player)
        first_attendance.overall = 80
        first_attendance.save(update_fields=["overall"])
        second_player = self.players[2]
        second_player.overall = 80
        second_player.save(update_fields=["overall"])
        second_attendance = MatchAttendance.objects.get(match=self.match, player=second_player)
        second_attendance.overall = 80
        second_attendance.save(update_fields=["overall"])
        second_rater = User.objects.create_user(
            username="comum-2",
            email="comum-2@pelada.local",
            password=self.password,
            role=Role.COMMON,
        )
        third_rater = User.objects.create_user(
            username="comum-3",
            email="comum-3@pelada.local",
            password=self.password,
            role=Role.COMMON,
        )
        self.match.status = Match.Status.ARCHIVED
        self.match.archived_at = timezone.now()
        self.match.save(update_fields=["status", "archived_at"])
        for rater in (self.common_user, second_rater, third_rater):
            MatchPlayerRating.objects.create(
                match=self.match,
                rater_user=rater,
                rated_attendance=first_attendance,
                rated_player=first_player,
                score=8,
            )
            MatchPlayerRating.objects.create(
                match=self.match,
                rater_user=rater,
                rated_attendance=second_attendance,
                rated_player=second_player,
                score=9,
            )

        response = self.client.post(f"/api/matches/{self.match.id}/finalize-ratings/")

        self.assertEqual(response.status_code, 200)
        first_player.refresh_from_db()
        second_player.refresh_from_db()
        self.assertEqual(first_player.overall, 81)
        self.assertEqual(second_player.overall, 85)

    def test_admin_can_recalculate_finalized_ratings_from_attendance_snapshot(self) -> None:
        target_player = self.players[1]
        target_player.overall = 92
        target_player.save(update_fields=["overall"])
        target_attendance = MatchAttendance.objects.get(match=self.match, player=target_player)
        target_attendance.overall = 90
        target_attendance.save(update_fields=["overall"])
        second_rater = User.objects.create_user(
            username="comum-2",
            email="comum-2@pelada.local",
            password=self.password,
            role=Role.COMMON,
        )
        third_rater = User.objects.create_user(
            username="comum-3",
            email="comum-3@pelada.local",
            password=self.password,
            role=Role.COMMON,
        )
        self.match.status = Match.Status.ARCHIVED
        self.match.archived_at = timezone.now()
        self.match.ratings_finalized_at = timezone.now()
        self.match.save(update_fields=["status", "archived_at", "ratings_finalized_at"])
        for rater in (self.common_user, second_rater, third_rater):
            MatchPlayerRating.objects.create(
                match=self.match,
                rater_user=rater,
                rated_attendance=target_attendance,
                rated_player=target_player,
                score=10,
            )

        response = self.client.post(f"/api/matches/{self.match.id}/recalculate-ratings/")

        self.assertEqual(response.status_code, 200)
        target_player.refresh_from_db()
        self.assertEqual(target_player.overall, 95)
        summary_entry = next(
            item for item in response.data["overall_summary"] if item["player_id"] == str(target_player.id)
        )
        self.assertEqual(summary_entry["previous_overall"], 90)
        self.assertEqual(summary_entry["current_overall"], 95)
        self.assertEqual(summary_entry["delta"], 5)

    def test_admin_can_create_transaction_via_api(self) -> None:
        response = self.client.post(
            "/api/transactions/",
            {
                "direction": Transaction.Direction.INFLOW,
                "category": Transaction.Category.EXTRA_FEE,
                "status": Transaction.Status.POSTED,
                "amount": "99.90",
                "description": "Taxa extra dos coletes",
                "occurred_on": date.today().isoformat(),
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(Transaction.objects.filter(description="Taxa extra dos coletes").count(), 1)

    def test_admin_can_create_player_with_manual_overall(self) -> None:
        response = self.client.post(
            "/api/players/",
            {
                "full_name": "Novo Linha",
                "preferred_position": Player.PreferredPosition.FORWARD,
                "monthly_fee_amount": "120.00",
                "overall": 82,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201)
        created_player = Player.objects.get(full_name="Novo Linha")
        self.assertEqual(created_player.overall, 82)

    def test_admin_can_update_position_and_overall(self) -> None:
        player = self.players[0]

        response = self.client.put(
            f"/api/players/{player.id}/",
            {
                "full_name": player.full_name,
                "nickname": player.nickname,
                "player_type": player.player_type,
                "preferred_position": Player.PreferredPosition.GOALKEEPER,
                "email": player.email,
                "phone_number": player.phone_number,
                "shirt_number": player.shirt_number,
                "monthly_fee_amount": str(player.monthly_fee_amount),
                "joined_on": player.joined_on,
                "is_active": player.is_active,
                "notes": player.notes,
                "overall": 87,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        player.refresh_from_db()
        self.assertEqual(player.preferred_position, Player.PreferredPosition.GOALKEEPER)
        self.assertEqual(player.overall, 87)

    def test_admin_can_update_transaction_and_link_player(self) -> None:
        transaction = Transaction.objects.filter(category=Transaction.Category.MONTHLY_FEE).first()
        assert transaction is not None

        response = self.client.put(
            f"/api/transactions/{transaction.id}/",
            {
                "direction": Transaction.Direction.INFLOW,
                "category": Transaction.Category.MONTHLY_FEE,
                "status": Transaction.Status.POSTED,
                "amount": "70.00",
                "description": "Mensalidade do Jogador 0",
                "occurred_on": date.today().isoformat(),
                "reference_month": date.today().replace(day=1).isoformat(),
                "related_player": str(self.players[0].id),
                "notes": "Pagamento confirmado no pix",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        transaction.refresh_from_db()
        self.assertEqual(transaction.related_player_id, self.players[0].id)
        self.assertEqual(transaction.amount, Decimal("70.00"))

    def test_admin_can_void_transaction_via_api(self) -> None:
        transaction = Transaction.objects.filter(category=Transaction.Category.FIELD_RENT).first()
        assert transaction is not None

        response = self.client.patch(
            f"/api/transactions/{transaction.id}/",
            {"status": Transaction.Status.VOIDED},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        transaction.refresh_from_db()
        self.assertEqual(transaction.status, Transaction.Status.VOIDED)
